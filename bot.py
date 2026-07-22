"""
bot.py — Telegram бот підбору сантехніки (Hotpoint).
Gemini OCR → keyword пошук → Claude вибір → Excel.

ЗНАЧКИ-КАРАКУЛІ МАЙСТРІВ (малюють замість слів!):
⊥ / Т-подібна вилка (навіть під кутом) = ТРІЙНИК: "⊥110х110х90" → Трійник канал 110х110 87.
∠ / L / дві з'єднані лінії кутом = КОЛІНО: "∠110х45" → Коліно канал 110 45.
-//- / —//— = ПОВТОР типу з попереднього рядка (новий розмір).
Рядок без типу ("х45=10шт", "ф50х1м=3шт") = продовження ПОПЕРЕДНЬОГО типу.

ПРОЕКТНІ ПОЗНАЧЕННЯ (специфікації PDF):
Ду15/Dn15=1/2 | Ду20=3/4 | Ду25=1 | Ду32=1 1/4 | Ду40=1 1/2 | Ду50=2
ЗР = зовнішня різьба (=МРЗ) | ВР = внутрішня (=МРВ)
Групи: "Труба Stabi Plus: Ekoplastik" + підрядки "32х4,4 м 40" → успадковуй тип
і виробника групи ("Труба PPR EVO ф32 Ekoplastik", 40 м). Виробник у колонці = виробник рядка.

⚠️ PUSH СТРОГО: push_systems ТІЛЬКИ якщо явно написано пуш/push/натяжн/PEX.
Голі розміри БЕЗ цих слів = НЕ push! "⊥110" = трійник каналізації, НЕ натяжний!
"""
import os, json, re, base64, threading, time
from io import BytesIO

import telebot
from flask import Flask, request as flask_request
from telebot.types import (Update, InlineKeyboardMarkup, InlineKeyboardButton,
                            ReplyKeyboardMarkup, KeyboardButton)
import anthropic
import pandas as pd
from google import genai as genai_new
from google.genai import types as genai_types

# ═══════════════════════════════════════════════════════════════════════════════
# ІНІЦІАЛІЗАЦІЯ
# ═══════════════════════════════════════════════════════════════════════════════
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "")
ANTHROPIC_KEY  = os.environ.get("ANTHROPIC_KEY", "")
GEMINI_KEY     = os.environ.get("GEMINI_KEY", "")

# КРИТИЧНО: threaded=False!
# За замовчуванням telebot виконує хендлери у своєму ThreadPool, де винятки
# зберігаються мовчки і піднімаються тільки в polling. З webhook вони ЗНИКАЮТЬ:
# process_new_updates повертається "успішно", а відповідь не відправляється.
_ExcBase = getattr(telebot, 'ExceptionHandler', object)

class _LogExc(_ExcBase):
    def handle(self, exception):
        import traceback
        print(f"❌ ПОМИЛКА В ХЕНДЛЕРІ: {exception}", flush=True)
        traceback.print_exc()
        return True

try:
    bot = telebot.TeleBot(TELEGRAM_TOKEN, threaded=False, exception_handler=_LogExc())
except TypeError:
    # стара версія telebot без exception_handler
    bot = telebot.TeleBot(TELEGRAM_TOKEN, threaded=False)

claude        = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
gemini_client = genai_new.Client(api_key=GEMINI_KEY)

# temperature=0 → однаковий вхід = однаковий вихід (прибирає рандом OCR)
try:
    _GEMCFG = genai_types.GenerateContentConfig(temperature=0)
except Exception:
    _GEMCFG = None

def _gemini_call(contents):
    kwargs = {"model": "gemini-2.5-flash", "contents": contents}
    if _GEMCFG is not None:
        kwargs["config"] = _GEMCFG
    return gemini_client.models.generate_content(**kwargs)

CATALOG_PATH = "catalog.json"
RULES_FILE   = "rules.txt"

# 💾 Постійне сховище (гілка GitHub botdata): відновлюємо дані ДО імпорту кешу
try:
    import storage
    storage.restore()
except Exception as _e:
    print(f"⚠️ storage restore: {_e}", flush=True)

# Кеш нормалізацій (auto/confirmed/banned) і профілі клієнтів
from cache import (cache_lookup, cache_save, cache_delete, get_cache,
                   cache_set_status, cache_ban_pair, is_banned as cache_is_banned)
import clients

# ═══════════════════════════════════════════════════════════════════════════════
# АДМІН
# ═══════════════════════════════════════════════════════════════════════════════
ADMIN_ID           = 395121797
PENDING_RULES_FILE = "pending_rules.json"
USAGE_LOG_FILE     = "usage_log.json"

def is_admin(uid: int) -> bool:
    return uid == ADMIN_ID

def log_usage(chat_id, username, total, found, files):
    try:
        log = []
        if os.path.exists(USAGE_LOG_FILE):
            with open(USAGE_LOG_FILE, encoding="utf-8") as f:
                log = json.load(f)
        log.append({"chat_id": chat_id, "username": username,
                    "date": time.strftime("%Y-%m-%d %H:%M"),
                    "total": total, "found": found, "files": files})
        log = log[-1000:]
        with open(USAGE_LOG_FILE, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False)
    except Exception as e:
        print(f"⚠️ log_usage: {e}")

def get_usage_stats() -> str:
    if not os.path.exists(USAGE_LOG_FILE):
        return "📊 Статистика порожня."
    try:
        with open(USAGE_LOG_FILE, encoding="utf-8") as f:
            log = json.load(f)
    except Exception:
        return "⚠️ Помилка читання логу."
    if not log:
        return "📊 Статистика порожня."
    users = {}
    for rec in log:
        u = rec.get("username", "?")
        if u not in users:
            users[u] = {"orders": 0, "total": 0, "found": 0}
        users[u]["orders"] += 1
        users[u]["total"]  += rec.get("total", 0)
        users[u]["found"]  += rec.get("found", 0)
    lines = [f"📊 *Статистика* ({len(log)} замовлень)\n"]
    for u, s in sorted(users.items(), key=lambda x: -x[1]["orders"]):
        pct = int(s["found"] / s["total"] * 100) if s["total"] else 0
        lines.append(f"👤 {u}: {s['orders']} замовл., {s['total']} поз., {pct}% знайдено")
    lines.append("\n🕐 Останні 5:")
    for rec in log[-5:]:
        lines.append(f"• {rec['date']} — {rec['username']}: {rec['found']}/{rec['total']}")
    return "\n".join(lines)

OCR_CORRECTIONS_FILE = "ocr_corrections.json"

def load_ocr_corrections() -> dict:
    """Словник OCR-корекцій: {неправильно: правильно}"""
    if os.path.exists(OCR_CORRECTIONS_FILE):
        try:
            with open(OCR_CORRECTIONS_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_ocr_correction(wrong: str, right: str):
    """Зберігає OCR-корекцію."""
    wrong = wrong.lower().strip()
    right = right.lower().strip()
    if not wrong or not right or wrong == right:
        return
    d = load_ocr_corrections()
    d[wrong] = right
    with open(OCR_CORRECTIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=2)

def get_ocr_prompt_block() -> str:
    """Блок корекцій для Gemini-промпту."""
    d = load_ocr_corrections()
    if not d:
        return ""
    lines = [f"  «{w}» часто насправді «{r}»" for w, r in d.items()]
    return ("\nЧАСТІ ПОМИЛКИ ЧИТАННЯ ЦЬОГО ПОЧЕРКУ (якщо бачиш ліве — "
            "придивись, дуже ймовірно це праве):\n" + "\n".join(lines))


NOT_FOUND_FILE = "not_found_log.json"

def log_not_found(rows: list):
    """Накопичує незнайдені позиції для звіту 'Діри каталогу'."""
    if not rows:
        return
    try:
        log = []
        if os.path.exists(NOT_FOUND_FILE):
            with open(NOT_FOUND_FILE, encoding="utf-8") as f:
                log = json.load(f)
        for r in rows:
            log.append({
                "original":   r.get("original", "")[:80],
                "normalized": r.get("normalized", "")[:80],
                "date":       time.strftime("%Y-%m-%d"),
            })
        log = log[-2000:]
        with open(NOT_FOUND_FILE, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False)
    except Exception as e:
        print(f"⚠️ log_not_found: {e}")

def get_catalog_gaps() -> str:
    """Топ незнайдених запитів — що треба додати в прайси."""
    if not os.path.exists(NOT_FOUND_FILE):
        return "🕳 Порожньо — все знаходилось."
    try:
        with open(NOT_FOUND_FILE, encoding="utf-8") as f:
            log = json.load(f)
    except Exception:
        return "⚠️ Помилка читання."
    if not log:
        return "🕳 Порожньо."
    groups = {}
    for rec in log:
        key = re.sub(r"\s+", " ", (rec.get("normalized") or rec.get("original","")).lower()).strip()
        if not key:
            continue
        if key not in groups:
            groups[key] = {"n": 0, "show": rec.get("normalized") or rec.get("original",""),
                           "orig": rec.get("original",""), "last": rec.get("date","")}
        groups[key]["n"] += 1
        groups[key]["last"] = rec.get("date", groups[key]["last"])
    top = sorted(groups.values(), key=lambda g: -g["n"])[:20]
    lines = [f"🕳 ДІРИ КАТАЛОГУ — топ незнайдених ({len(log)} записів):\n"]
    for i, g in enumerate(top, 1):
        lines.append(f"{i}. ×{g['n']}  {g['show'][:48]}")
        if g['orig'] and g['orig'].lower() != g['show'].lower():
            lines.append(f"      (писали: {g['orig'][:45]})")
    lines.append("\n➡️ Ці товари варто додати в прайси або створити правило.")
    return "\n".join(lines)


def load_pending_rules():
    if os.path.exists(PENDING_RULES_FILE):
        try:
            with open(PENDING_RULES_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return []

def save_pending_rules(rules):
    with open(PENDING_RULES_FILE, "w", encoding="utf-8") as f:
        json.dump(rules, f, ensure_ascii=False, indent=2)

# ─── Черга ВИПРАВЛЕНЬ від звичайних користувачів (адмін підтверджує) ─────────
PENDING_FIXES_FILE = "pending_fixes.json"

def load_pending_fixes() -> list:
    if os.path.exists(PENDING_FIXES_FILE):
        try:
            with open(PENDING_FIXES_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return []

def save_pending_fixes(fixes: list):
    with open(PENDING_FIXES_FILE, "w", encoding="utf-8") as f:
        json.dump(fixes, f, ensure_ascii=False, indent=2)

def add_pending_fix(fix: dict) -> int:
    fixes = load_pending_fixes()
    fix["date"] = time.strftime("%Y-%m-%d %H:%M")
    fixes.append(fix)
    save_pending_fixes(fixes)
    return len(fixes)

def apply_fix(fix: dict):
    """Застосовує підтверджене виправлення: старе → banned, нове → confirmed."""
    original = fix.get('original', '')
    cat      = fix.get('category', 'other')
    old      = fix.get('old_name')
    new      = fix.get('new_name')
    slug     = fix.get('client_slug')
    if old:
        cache_ban_pair(original, old, cat)
        if slug:
            clients.client_cache_set_status(slug, original, old, 'banned')
    if new:
        cache_save(original, {}, fix.get('normalized', original), new, cat, 100)
        cache_set_status(original, new, 'confirmed')
        if slug:
            clients.client_cache_save(slug, original, new, cat, 100)
            clients.client_cache_set_status(slug, original, new, 'confirmed')

_kn_pending = {}   # chat_id → згенероване правило що чекає ✅/❌

def suggest_knowledge_rule(chat_id, original, old_name, new_name):
    """
    Після навчання бот САМ формулює правило для бази знань (rules.txt)
    і пропонує адміну додати одним тапом. База знань росте розмовою.
    """
    try:
        prompt = (f"Менеджер-сантехнік виправив підбір товару.\n"
                  f"Написано в замовленні: «{original}»\n"
                  f"Бот вибрав (НЕПРАВИЛЬНО): «{old_name or '(не знайшов)'}»\n"
                  f"Правильна відповідь: «{new_name}»\n\n"
                  f"Сформулюй ОДНЕ коротке правило (до 15 слів) українською, яке допоможе "
                  f"боту наступного разу зрозуміти такий запит правильно. Правило має бути "
                  f"загальним (про термін/скорочення/тип), а не про цей конкретний рядок.\n"
                  f"Якщо корисного загального правила сформулювати не можна — напиши SKIP.\n"
                  f"Відповідь: тільки текст правила або SKIP.")
        resp = claude.messages.create(model="claude-sonnet-4-5", max_tokens=100,
                                      messages=[{"role": "user", "content": prompt}])
        rule = resp.content[0].text.strip().strip('"«»')
        if not rule or 'SKIP' in rule.upper() or len(rule) > 200:
            return
        _kn_pending[chat_id] = rule
        mk = InlineKeyboardMarkup()
        mk.add(InlineKeyboardButton("✅ Додати в базу знань", callback_data="knok"),
               InlineKeyboardButton("❌ Ні", callback_data="knno"))
        bot.send_message(chat_id,
            f"💡 Бот пропонує нове правило з цього виправлення:\n\n_{rule}_\n\n"
            f"Додати? (потрапить у знання для всіх наступних розпізнавань)",
            parse_mode="Markdown", reply_markup=mk)
    except Exception as e:
        print(f"⚠️ suggest_rule: {e}")


@bot.callback_query_handler(func=lambda c: c.data.startswith("ocrs_"))
def handle_ocr_pair_save(call):
    st = _train_state.get(call.message.chat.id)
    pairs = (st or {}).get('ocr_pairs') or []
    idx = int(call.data[5:])
    if idx >= len(pairs):
        bot.answer_callback_query(call.id, "Застаріло"); return
    w, rt = pairs[idx]
    save_ocr_correction(w, rt)
    bot.answer_callback_query(call.id, f"Збережено: {w}→{rt}")
    try:
        bot.edit_message_text(f"✅ Корекція почерку збережена: «{w}» → «{rt}»",
            call.message.chat.id, call.message.message_id)
    except Exception:
        pass


@bot.callback_query_handler(func=lambda c: c.data in ("knok", "knno"))
def handle_knowledge_decision(call):
    rule = _kn_pending.pop(call.message.chat.id, None)
    if call.data == "knok" and rule:
        add_rule(rule)
        try:
            bot.edit_message_text(f"✅ Додано в базу знань:\n_{rule}_",
                call.message.chat.id, call.message.message_id, parse_mode="Markdown")
        except Exception: pass
        bot.answer_callback_query(call.id, "Збережено")
    else:
        try:
            bot.edit_message_text("❌ Пропущено.", call.message.chat.id, call.message.message_id)
        except Exception: pass
        bot.answer_callback_query(call.id)


def notify_admin_fix(username, original, old_name, new_name, n):
    try:
        bot.send_message(ADMIN_ID,
            f"🔔 Виправлення від @{username} (в черзі: {n})\n"
            f"«{(original or '')[:45]}»\n"
            f"❌ {(old_name or '—')[:55]}\n"
            f"✅ {(new_name or '(тільки заборонити старе)')[:55]}\n\n"
            f"Підтвердити: 👑 Правила на розгляд")
    except Exception:
        pass

# ═══════════════════════════════════════════════════════════════════════════════
# КАТАЛОГ
# ═══════════════════════════════════════════════════════════════════════════════
CATALOG_FILES = [
    ("adapters_reducers","adapters_reducers"),("automation","automation"),
    ("boilers","boilers"),("fasteners_sealants","fasteners_sealants"),
    ("filtration","filtration"),("heating","heating"),("hoses","hoses"),
    ("insulation","insulation"),("metal_plastic","metal_plastic"),
    ("mixers_faucets","mixers_faucets"),("plastic_ppr","plastic_ppr"),
    ("pumps","pumps"),("push_systems","push_systems"),
    ("radiators_radiatorsvalve","radiators_radiatorsvalve"),
    ("safety_valves","safety_valves"),("sanitary_ware","sanitary_ware"),
    ("sewage","sewage"),("shutoff_valves","shutoff_valves"),
    ("siphons_fittings","siphons_fittings"),("towel_warmers","towel_warmers"),
    ("underfloor_heating","underfloor_heating"),("water_heaters","water_heaters"),
    ("water_meters","water_meters"),
]

def is_header_row(name, artikul, or_val):
    name = str(name).strip()
    if not name or name == 'nan':
        return True
    art = str(artikul).strip()
    if art and art not in ('nan', '0', ''):
        return False
    try:
        if float(or_val) > 0:
            return False
    except Exception:
        pass
    return True

def build_catalog_from_xlsx():
    catalog = []
    bot_dir = os.path.dirname(os.path.abspath(__file__)) if '__file__' in globals() else os.getcwd()
    search_dirs = ['.', 'src', bot_dir]
    for key, category in CATALOG_FILES:
        found = False
        for d in search_dirs:
            path = os.path.join(d, f"{key}.xlsx")
            if not os.path.exists(path):
                continue
            try:
                df = pd.read_excel(path, header=0)
                cols = list(df.columns)
                rename = {}
                if len(cols) >= 1: rename[cols[0]] = 'name'
                if len(cols) >= 2: rename[cols[1]] = 'artikul'
                if len(cols) >= 3: rename[cols[2]] = 'price'
                df = df.rename(columns=rename)
                count = 0
                for _, row in df.iterrows():
                    name    = str(row.get('name', '')).strip()
                    artikul = row.get('artikul', '')
                    price   = row.get('price', 0)
                    if is_header_row(name, artikul, price):
                        continue
                    try:
                        p = float(price)
                    except Exception:
                        p = 0.0
                    art = str(artikul).strip()
                    name_full = name
                    name_clean = re.sub(r'\s*\{[^}]+\}', '', name).strip()
                    catalog.append({
                        'name': name_clean, 'name_full': name_full,
                        'artikul': art if art != 'nan' else '',
                        'category': category, 'price': p,
                    })
                    count += 1
                print(f"  ✅ {key}: {count} товарів")
                found = True
                break
            except Exception as e:
                print(f"  ❌ {key}: {e}")
        if not found:
            print(f"  ⚠️ {key}.xlsx не знайдено")
    return catalog

print("📦 Завантажую каталог...", flush=True)
if os.path.exists(CATALOG_PATH):
    with open(CATALOG_PATH, encoding="utf-8") as f:
        CATALOG = json.load(f)
    print(f"✅ Каталог: {len(CATALOG)} позицій", flush=True)
else:
    print("⚙️ Будую catalog.json...", flush=True)
    CATALOG = build_catalog_from_xlsx()
    if CATALOG:
        with open(CATALOG_PATH, "w", encoding="utf-8") as f:
            json.dump(CATALOG, f, ensure_ascii=False)
        print(f"✅ Збережено: {len(CATALOG)} позицій", flush=True)

# Видаляємо виведені
before = len(CATALOG)
CATALOG = [c for c in CATALOG if 'виведено з асортименту' not in c.get('name','').lower()]
if len(CATALOG) < before:
    print(f"🧹 Видалено {before - len(CATALOG)} виведених", flush=True)

# Індексуємо токени (lazy)
_tokens_built = False

def tokenize(text: str) -> set:
    t = text.lower()
    t = re.sub(r'(\d+)/(\d+)', r'\1_\2', t)
    t = re.sub(r'[фfдd]\s*(\d)', r'\1', t)
    t = re.sub(r'(\d+)\s*[хxX×]\s*(\d+)/(\d+)', r'\1 \2_\3', t)
    def strip_thick(m):
        d1, d2 = m.group(1), m.group(2)
        if '_' in d2: return m.group(0)
        try:
            if float(d2.replace(',', '.')) < 15: return d1
        except Exception: pass
        return m.group(0)
    t = re.sub(r'(\d+)\s*[хxX×]\s*(\d+(?:[.,]\d+)?)(?![\d_])', strip_thick, t)
    t = re.sub(r'(8[67])[.,]5', r'\1', t)
    t = re.sub(r'(?<![0-9_])\d+[.,]\d+(?![0-9_])', '', t)
    return set(re.findall(r'[а-яёіїєґa-z]+|[0-9]+_[0-9]+|[0-9]+', t))

def ensure_tokens():
    global _tokens_built
    if not _tokens_built:
        print("🔨 Індексую токени...", flush=True)
        for item in CATALOG:
            item['_tokens'] = tokenize(item['name'])
            item['_attrs'] = parse_attrs(item['name'])
        print("✅ Індексація завершена", flush=True)
        _tokens_built = True

# ═══════════════════════════════════════════════════════════════════════════════
# ПРАВИЛА
# ═══════════════════════════════════════════════════════════════════════════════
def get_rules() -> str:
    if not os.path.exists(RULES_FILE): return ""
    with open(RULES_FILE, encoding="utf-8") as f: return f.read().strip()

def add_rule(rule: str):
    with open(RULES_FILE, "a", encoding="utf-8") as f:
        f.write(f"- {rule}\n")

# ═══════════════════════════════════════════════════════════════════════════════
# ПОШУК
# ═══════════════════════════════════════════════════════════════════════════════
BRAND_TOKENS = {
    'raftec':      ['raftec', 'RAFTEC'],
    'рафтек':      ['raftec', 'RAFTEC'],
    'ekoplastik':  ['ekoplastik', 'Ekoplastik'],
    'екопластик':  ['ekoplastik', 'Ekoplastik'],
    'екопласт':    ['ekoplastik', 'Ekoplastik'],
    'asg':         ['asg', 'ASG'],
    'асг':         ['asg', 'ASG'],
    'ostendorf':   ['ostendorf', 'OSTENDORF'],
    'остендорф':   ['ostendorf', 'OSTENDORF'],
    'plm':         ['plm', 'PLM'],
    'плм':         ['plm', 'PLM'],
    'hidros':      ['hidros', 'Hidros', 'HIDROS'],
    'хідрос':      ['hidros', 'Hidros'],
    'гідрос':      ['hidros', 'Hidros'],
    'idmar':       ['idmar', 'IDMAR'],
    'termojet':    ['termojet', 'Termojet'],
    'термоджет':   ['termojet', 'Termojet'],
    'tatra':       ['tatra', 'TATRA', 'Tatra-Line'],
    'татра':       ['tatra', 'TATRA'],
    'rehau':       ['rehau', 'REHAU'],
    'рехау':       ['rehau', 'REHAU'],
    'ecosoft':     ['ecosoft', 'Ecosoft'],
    'екософт':     ['ecosoft', 'Ecosoft'],
    'biasi':       ['biasi', 'BIASI'],
    'біасі':       ['biasi', 'BIASI'],
    'valrom':      ['valrom', 'Valrom'],
    'unipak':      ['unipak', 'Unipak'],
    'kan':         ['kan', 'KAN'],
    'herz':        ['herz', 'HERZ', 'Herz'],
    'герц':        ['herz', 'HERZ'],
    'giacomini':   ['giacomini', 'Giacomini'],
    'джикоміні':   ['giacomini', 'Giacomini'],
    'danfoss':     ['danfoss', 'Danfoss'],
    'данфос':      ['danfoss', 'Danfoss'],
    'wilo':        ['wilo', 'WILO'],
    'віло':        ['wilo', 'WILO'],
    'grundfos':    ['grundfos', 'GRUNDFOS'],
    'грундфос':    ['grundfos', 'GRUNDFOS'],
    'bonomi':      ['bonomi', 'Bonomi'],
    'бономі':      ['bonomi', 'Bonomi'],
    'pattaroni':   ['pattaroni', 'Pattaroni'],
    'k-flex':      ['k-flex', 'K-FLEX'],
    'kflex':       ['k-flex', 'K-FLEX'],
    'кфлекс':      ['k-flex', 'K-FLEX'],
    'valsir':      ['valsir', 'Valsir'],
    'meibes':      ['meibes', 'Meibes'],
    'flamco':      ['flamco', 'Flamco'],
    'reflex':      ['reflex', 'Reflex'],
    'icma':        ['icma', 'Icma'],
    'drazice':     ['drazice', 'Drazice'],
    'дражице':     ['drazice', 'Drazice'],
    'vaillant':    ['vaillant', 'Vaillant'],
    'вайлант':     ['vaillant', 'Vaillant'],
    'alcaplast':   ['alcaplast', 'AlcaPlast'],
    'esbe':        ['esbe', 'ESBE'],
    'grohe':       ['grohe', 'Grohe'],
    'thermaflex':  ['thermaflex', 'Thermaflex'],
}

# Категорії — те що менеджер може написати в підказці
CATEGORY_ALIASES = {
    'каналізація': 'sewage', 'канал': 'sewage', 'каналізац': 'sewage',
    'пайка': 'plastic_ppr', 'ппр': 'plastic_ppr', 'ppr': 'plastic_ppr',
    'пластик': 'plastic_ppr', 'поліпропілен': 'plastic_ppr',
    'кран': 'shutoff_valves', 'крани': 'shutoff_valves',
    'арматура': 'shutoff_valves', 'вентил': 'shutoff_valves',
    'пуш': 'push_systems', 'push': 'push_systems', 'пекс': 'push_systems',
    'pex': 'push_systems', 'натяжн': 'push_systems', 'гільз': 'push_systems',
    'насос': 'pumps', 'насоси': 'pumps',
    'радіатор': 'radiators_radiatorsvalve', 'радіатори': 'radiators_radiatorsvalve',
    'утепл': 'insulation', 'ізоляц': 'insulation', 'мірелон': 'insulation',
    'фільтр': 'filtration', 'очист': 'filtration',
    'металопласт': 'metal_plastic', 'мп': 'metal_plastic',
    'котел': 'boilers', 'котли': 'boilers',
    'бойлер': 'water_heaters', 'водонагрів': 'water_heaters',
    'тепла підлога': 'underfloor_heating', 'тп': 'underfloor_heating',
    'сифон': 'siphons_fittings',
    'змішувач': 'mixers_faucets',
    'кріплення': 'fasteners_sealants', 'хомут': 'fasteners_sealants',
    'перехідник': 'adapters_reducers',
    'опалення': 'heating',
    'шланг': 'hoses',
    'рушникосуш': 'towel_warmers',
}

DEFAULT_BRAND_PRIORITY = {
    'sewage':         [['asg','ASG'], ['ostendorf','OSTENDORF']],
    'plastic_ppr':    [['ekoplastik','Ekoplastik'], ['asg','ASG'], ['raftec','RAFTEC']],
    'shutoff_valves': [['raftec','RAFTEC']],
    'adapters_reducers': [['raftec','RAFTEC']],
    'filtration':     [['ecosoft','Ecosoft']],
    'radiators_radiatorsvalve': [['hidros','Hidros'], ['idmar','IDMAR']],
    'pumps':          [['tatra','TATRA'], ['termojet','Termojet']],
    'insulation':     [['plm','PLM']],
    'push_systems':   [['raftec','RAFTEC'], ['rehau','REHAU']],
    'metal_plastic':  [['raftec','RAFTEC']],
    'fasteners_sealants': [['eco','ECO']],
}

# ═══════════════════════════════════════════════════════════════════════════════
# АТРИБУТНИЙ ПОШУК: тип + діаметри + кут + різьба (детермінований, точний)
# Побудовано на аналізі всіх 48 879 позицій каталогу з GitHub.
# ═══════════════════════════════════════════════════════════════════════════════

# Сленг/синоніми → канонічний тип (обидві сторони зводяться до одного)
TYPE_SYNONYMS = {
    # коліно
    'коліно': 'коліно', 'колено': 'коліно', 'кутник': 'коліно', 'кут': 'коліно',
    'угол': 'коліно', 'уголок': 'коліно', 'відвід': 'коліно', 'отвод': 'коліно',
    'відведення': 'відведення',  # спец-фітинг з кількома кутами — ОКРЕМИЙ тип!
    # трійник
    'трійник': 'трійник', 'тройник': 'трійник', 'трійники': 'трійник',
    # труба
    'труба': 'труба', 'труби': 'труба',
    # муфта / з'єднання
    'муфта': 'муфта', "з'єднувач": 'муфта', 'зєднувач': 'муфта', 'соединитель': 'муфта',
    "з'єднання": 'американка', 'американка': 'американка', 'американки': 'американка',
    # кран
    'кран': 'кран', 'крани': 'кран', 'вентиль': 'кран', 'вентель': 'кран',
    # перехід
    'перехід': 'перехід', 'переход': 'перехід', 'перехідник': 'перехід',
    'редукція': 'перехід', 'редукци': 'перехід', 'футорка': 'футорка',
    # інші фітинги
    'заглушка': 'заглушка', 'хрестовина': 'хрестовина', 'крестовина': 'хрестовина',
    'ревізія': 'ревізія', 'ревизия': 'ревізія',
    'ніпель': 'ніпель', 'нипель': 'ніпель', 'штуцер': 'штуцер',
    'подовжувач': 'подовжувач', 'бочонок': 'подовжувач', 'бочка': 'подовжувач',
    'згін': 'згін', 'згон': 'згін', 'напівзгін': 'напівзгін',
    'гільза': 'гільза', 'кільце': 'кільце',  # НЕ синоніми! (перевірено по каталогу)
    'фланець': 'фланець',
    # обладнання
    'клапан': 'клапан', 'засувка': 'засувка', 'затвор': 'затвор',
    'фільтр': 'фільтр', 'фильтр': 'фільтр',
    'насос': 'насос', 'радіатор': 'радіатор', 'радиатор': 'радіатор',
    'колектор': 'колектор', 'гребінка': 'колектор', 'гребенка': 'колектор',
    'термоголовка': 'термоголовка', 'сифон': 'сифон', 'трап': 'трап',
    'хомут': 'хомут', 'опора': 'опора', 'скоба': 'скоба',
    'утеплювач': 'утеплювач', 'мірелон': 'утеплювач', 'мирелон': 'утеплювач',
    'шланг': 'шланг', 'підводка': 'підводка', 'підведення': 'підводка',
    'котел': 'котел', 'бойлер': 'водонагрівач', 'водонагрівач': 'водонагрівач',
    'змішувач': 'змішувач', 'мийка': 'мийка', 'умивальник': 'умивальник',
    'лічильник': 'лічильник', 'счетчик': 'лічильник',
    'стрічка': 'стрічка', 'шпилька': 'шпилька', 'дюбель': 'дюбель',
    'вузол': 'вузол', 'комплект': 'комплект', 'набір': 'комплект',
    'шафа': 'шафа', 'бак': 'бак', 'ємність': 'бак', 'емкость': 'бак',
    'группа': 'група', 'група': 'група',
}
# Сусідні категорії (Gemini міг промахнутись групою — шукаємо і там)
SIMILAR_CATS = {
    'plastic_ppr':      ['adapters_reducers', 'heating'],
    'adapters_reducers':['plastic_ppr', 'shutoff_valves'],
    'heating':          ['plastic_ppr', 'shutoff_valves'],
    'push_systems':     ['metal_plastic', 'plastic_ppr'],
    'metal_plastic':    ['push_systems', 'adapters_reducers'],
    'sewage':           ['siphons_fittings'],
    'siphons_fittings': ['sewage'],
    'shutoff_valves':   ['adapters_reducers', 'safety_valves'],
    'underfloor_heating':['metal_plastic', 'push_systems'],
    'insulation':       ['fasteners_sealants'],
    'fasteners_sealants':['insulation'],
}

# ПРАВИЛА КУТІВ ПО СИСТЕМАХ (від Тараса):
#   каналізація: виріб 87°, майстри пишуть 90 → нормалізуємо 90→87
#   пайка PPR:   тільки 45 і 90 (як є)
#   PUSH:        тільки 90, у назвах кута НЕМАЄ → ігноруємо кут взагалі
def normalize_angle_for_category(angle, category):
    if angle is None:
        return None
    if category == 'sewage' and angle == 90:
        return 87
    if category == 'push_systems':
        return None
    return angle

# Regex: довші ключі перші (щоб "напівзгін" не з'їдався "згін")
_TYPE_RE = re.compile(
    r'(?<![а-яёіїєґa-z])(' +
    '|'.join(sorted((re.escape(k) for k in TYPE_SYNONYMS), key=len, reverse=True)) +
    r')(?![а-яёіїєґ])', re.IGNORECASE)


def parse_attrs(text: str) -> dict:
    """
    Витягує атрибути з назви/запиту:
      type: канонічний тип ('коліно', 'трійник'...)
      dia:  список діаметрів [110, 50]
      angle: кут (87, 45...; 90 нормалізується окремо при матчі)
      thread: різьба '1_2', '3_4'...
    """
    t = text.lower()
    # Тип — перший знайдений
    typ = None
    m = _TYPE_RE.search(t)
    if m:
        typ = TYPE_SYNONYMS.get(m.group(1).lower())

    # Кут: "х 87,5°" / "x45°" / "87 град"
    angle = None
    ma = re.search(r'[хx]\s*(15|30|45|67|87|90)(?:[.,]5)?\s*[°º]', t)
    if not ma:
        ma = re.search(r'(?<![0-9])(15|30|45|67|87|90)(?:[.,]5)?\s*(?:[°º]|град)', t)
    if not ma:
        # Без °: "на 90", "х45" (тільки кути ≥45 щоб не плутати з розмірами)
        ma = re.search(r'(?:\bна\s+|[хx×]\s*)(45|67|87|90)(?:[.,]5)?(?![\d_.,])', t)
    if ma:
        angle = int(ma.group(1))

    # Різьба: токенайзер вже дає N_M
    toks = tokenize(text)
    thread = next((tk for tk in toks if '_' in tk), None)

    # Діаметри: числові токени 10-630 (реальний діапазон каталогу), без кута
    dias = []
    for tk in toks:
        if tk.isdigit():
            v = int(tk)
            if 10 <= v <= 630 and v != angle:
                dias.append(v)
    # Кут міг потрапити і як діаметр (напр. 45 і кут 45) — якщо кут є і в dias двічі, ок
    return {'type': typ, 'dia': sorted(set(dias)), 'angle': angle, 'thread': thread}


def _angle_match(qa_angle, item_angle) -> bool:
    """Строгий матч кута (нормалізація 90→87 для каналізації — у build_qa)."""
    if qa_angle is None:
        return True
    return item_angle == qa_angle


def attr_search(qa: dict, top_n: int = 10, brand_tokens: list = None, category: str = None) -> list[dict]:
    """
    Детермінований пошук по атрибутах. Tier-и (перший непорожній перемагає):
      1: тип + всі діаметри точно (set==) + кут          → 100%
      2: тип + діаметри запиту ⊆ кандидата + кут          → 95%
      3: тип + діаметри ⊆                                 → 85%
      4: діаметри ⊆ (без типу — коли тип не розпізнано)  → 70%
    Порожньо → викликаючий падає на старий keyword_search.
    """
    ensure_tokens()
    if not qa.get('dia') and not qa.get('type'):
        return []
    brand_lc = [t.lower() for t in brand_tokens] if brand_tokens else None
    q_dia = set(qa.get('dia') or [])
    q_type = qa.get('type')
    q_thread = qa.get('thread')

    # ІЄРАРХІЯ ПОШУКУ: 1) група товару → 2) виробник → 3) тип → 4) діаметри → 5) кут
    # Пули: спершу своя категорія, порожньо → сусідні, порожньо → весь каталог
    if category:
        pools = [
            [it for it in CATALOG if it.get('category') == category],
            [it for it in CATALOG if it.get('category') in SIMILAR_CATS.get(category, [])],
            CATALOG,
        ]
    else:
        pools = [CATALOG]

    for pool in pools:
        tiers = {1: [], 2: [], 3: [], 4: []}
        for item in pool:
            if brand_lc and not any(t in item['name'].lower() for t in brand_lc):
                continue
            ia = item.get('_attrs')
            if ia is None:
                continue
            i_dia = set(ia['dia'])
            # Різьба: якщо в запиті є — має бути і в кандидаті
            if q_thread and ia['thread'] != q_thread:
                continue
            type_ok = (q_type is not None and ia['type'] == q_type)
            dia_sub = bool(q_dia) and q_dia.issubset(i_dia)
            dia_eq  = dia_sub and (q_dia == i_dia)
            ang_ok  = _angle_match(qa.get('angle'), ia['angle'])

            if type_ok and dia_eq and ang_ok:
                tiers[1].append(item)
            elif type_ok and dia_sub and ang_ok:
                tiers[2].append(item)
            elif type_ok and dia_sub:
                tiers[3].append(item)
            elif dia_sub and q_type is None:
                tiers[4].append(item)

        _found_in_pool = False
        for tier, pct in ((1, 100), (2, 95), (3, 85), (4, 70)):
            cand = tiers[tier]
            if not cand:
                continue
            _found_in_pool = True
            # Ранжування: менше зайвих діаметрів; серії-слова; (п/з) вниз
            q_series = {w for w in tokenize(qa.get('_raw','')) if not w.isdigit() and '_' not in w}
            def score(it):
                ia = it['_attrs']
                extra_dia = len(set(ia['dia']) - q_dia)
                series_hit = len(q_series & it.get('_tokens', set()))
                pz = 5 if '(п/з)' in it['name'] else 0
                return (-series_hit, extra_dia, pz, len(it['name']))
            cand.sort(key=score)
            out = []
            for it in cand[:top_n]:
                c = dict(it)
                c['_match_pct'] = pct
                c['_attr_tier'] = tier
                out.append(c)
            return out
        # цей пул порожній у всіх tier → пробуємо наступний (сусідні/весь каталог)
    return []


def build_qa(пос: dict) -> dict:
    """Атрибути запиту: Gemini-поля + парсинг normalized + original (merge)."""
    qa = {'type': None, 'dia': [], 'angle': None, 'thread': None}
    # 1) Прямо від Gemini (найточніше — він бачив фото)
    g_dia = пос.get('dia')
    if isinstance(g_dia, list):
        qa['dia'] = [int(x) for x in g_dia if str(x).isdigit()]
    if пос.get('type'):
        qa['type'] = TYPE_SYNONYMS.get(str(пос['type']).lower().strip())
    if пос.get('angle') not in (None, '', 0):
        try: qa['angle'] = int(пос['angle'])
        except Exception: pass
    if пос.get('thread'):
        qa['thread'] = str(пос['thread']).replace('/', '_').strip()
    # 2) Доповнюємо парсингом текстів
    for txt in (пос.get('normalized',''), пос.get('original','')):
        if not txt: continue
        pa = parse_attrs(txt)
        if not qa['type']: qa['type'] = pa['type']
        if not qa['dia']: qa['dia'] = pa['dia']
        if qa['angle'] is None: qa['angle'] = pa['angle']
        if not qa['thread']: qa['thread'] = pa['thread']
    qa['_raw'] = f"{пос.get('normalized','')} {пос.get('original','')}"
    # Правила кутів по системах
    qa['angle'] = normalize_angle_for_category(qa['angle'], пос.get('category'))
    return qa


def smart_search(пос: dict, top_n: int = 12, brand_tokens: list = None) -> list[dict]:
    """Атрибутний пошук → fallback на старий keyword_search."""
    qa = пос.get('_qa')
    if qa is None:
        qa = build_qa(пос)
        пос['_qa'] = qa
    cand = attr_search(qa, top_n=top_n, brand_tokens=brand_tokens,
                       category=пос.get('category'))
    if cand:
        return cand
    return keyword_search(пос.get('normalized','') or пос.get('original',''),
                          top_n=top_n, brand_tokens=brand_tokens)


def validate_pick(qa: dict, item: dict) -> bool:
    """Пост-валідація вибору: діаметри запиту мусять бути в кандидаті; тип збігається."""
    ia = item.get('_attrs') or parse_attrs(item.get('name',''))
    q_dia = set(qa.get('dia') or [])
    if q_dia and not q_dia.issubset(set(ia['dia'])):
        return False
    if qa.get('type') and ia.get('type') and qa['type'] != ia['type']:
        # виняток: відведення ≈ коліно (обидва — повороти)
        if {qa['type'], ia['type']} != {'коліно', 'відведення'}:
            return False
    if qa.get('angle') and not _angle_match(qa['angle'], ia.get('angle')):
        return False
    return True


def keyword_search(query: str, top_n: int = 12, brand_tokens: list = None) -> list[dict]:
    ensure_tokens()
    q_tokens  = tokenize(query)
    q_numbers = set(re.findall(r'\d+', query.lower()))
    q_words   = q_tokens - q_numbers
    if not q_tokens: return []

    brand_lc = [t.lower() for t in brand_tokens] if brand_tokens else None
    scores = []
    for item in CATALOG:
        if brand_lc and not any(t in item['name'].lower() for t in brand_lc):
            continue
        it = item.get('_tokens', set())
        num_hits  = len(q_numbers & it)
        word_hits = len(q_words & it)
        if num_hits == 0 and word_hits == 0: continue
        raw = num_hits * 3 + word_hits
        penalty = max(0, len(it) - len(q_tokens)) * 0.1
        if '(п/з)' in item['name']: penalty += 5
        pct = min(int(raw / max(len(q_numbers)*3 + len(q_words), 1) * 100), 100)
        scores.append((raw - penalty, pct, item))
    scores.sort(key=lambda x: -x[0])
    result = []
    for _, pct, item in scores[:top_n]:
        c = dict(item)
        c['_match_pct'] = pct
        result.append(c)
    return result

def parse_caption_brands(caption: str) -> dict:
    """
    ПОСТРОКОВИЙ парсинг (кожен рядок/кома = окрема пара категорія+виробник).
    Фікс: бренд з попереднього рядка більше НЕ краде чужу категорію.
    """
    if not caption or not caption.strip():
        return {}
    result = {}
    chunks = re.split(r'[\n,;|]+', caption.lower())

    def _find_brands(text):
        out = {}
        for bk, bt in BRAND_TOKENS.items():
            for m in re.finditer(r'(?<![a-zа-яёіїєґ0-9])' + re.escape(bk) + r'(?![a-zа-яёіїєґ0-9])', text):
                out[m.start()] = bt
        return out

    def _find_cats(text):
        out = {}
        for alias, cat in CATEGORY_ALIASES.items():
            for m in re.finditer(r'(?<![a-zа-яёіїєґ])' + re.escape(alias), text):
                out[m.start()] = cat
        return out

    all_text = caption.lower()
    if re.search(r'(?<![а-я])(усе|все|всё|all)(?![а-я])', all_text):
        fb = _find_brands(all_text)
        if fb:
            first = fb[min(fb)]
            for cat in set(CATEGORY_ALIASES.values()):
                result[cat] = first
            return result

    global_brands = {}
    for chunk in chunks:
        chunk = chunk.strip()
        if not chunk:
            continue
        brands = _find_brands(chunk)
        cats = _find_cats(chunk)
        if brands and cats:
            for cpos, cat in cats.items():
                best, best_score = None, 1e9
                for bpos, bt in brands.items():
                    d = bpos - cpos
                    score = d if d >= 0 else abs(d) + 100
                    if score < best_score:
                        best_score, best = score, bt
                if best and cat not in result:
                    result[cat] = best
        elif brands and not cats:
            global_brands.update(brands)

    if not result and global_brands:
        first = global_brands[min(global_brands)]
        for cat in set(CATEGORY_ALIASES.values()):
            result[cat] = first
    return result


# ═══ ЗНАННЯ САНТЕХНІКИ (для Gemini) ════════════════════════════════════════
ЗНАННЯ = """
АБРЕВІАТУРИ (PPR): МРЗ=зовн.різьба; МРВ=внутр.; КРЗ=коліно зовн.; КРВ=внутр.
2 параметри (25х3/4)=муфта/коліно з різьбою; 3 параметри (20х16х20)=трійник. Кран ВВ→МРЗ; ВЗ→МРВ.

НОРМАЛІЗАЦІЯ — КОРОТКО (не вигадуй товщину/PN!):
"Труба PPR Fiber ф25 RAFTEC" | "Коліно канал ф110 87 OSTENDORF" | "Муфта PPR МРЗ 25х3/4 Ekoplastik"
Виробника пиши ТІЛЬКИ якщо він у підказці менеджера або в рядку; інакше НЕ додавай!

ТРУБИ PPR: опалення→Fiber(RAFTEC)/Faser(ASG)/EVO(Ekoplastik); холодна→PN20.
⚠️ Ekoplastik = серія EVO (Fiber Basalt/STABI виведено)!
Редукція PPR: Ekoplastik="Муфта перехідна PPR ВВ"; RAFTEC="Муфта редукційна". Шрабер=НЕ фітинг!

КАНАЛІЗАЦІЯ: 90→пиши "87" (без кому!). ASG=HTR; OSTENDORF=HT Safe.
ф40=умивальник, ф50=ванна/душ, ф110=стояк/унітаз. Компенсаційна муфта="Муфта вставна".

⚠️ PUSH МАРКЕРИ: пуш/push/натяжн/PEX/пекс/ГІЛЬЗ (слово "гільзи" в списку = ВСЕ замовлення PUSH,
бо гільзи існують тільки в натяжному монтажі!). Голі розміри без маркерів = НЕ push.
⚠️ НЕ додавай гільзи до фітингів сам — бот порахує їх автоматично.
⚠️ "труба 25 - 40м + ізол" — збережи "+ ізол" в original, НЕ створюй позицій утеплювача:
бот сам додасть синій+червоний PLM по половині метражу.
PUSH16≈PPR20. "Кутник натяжний ф25 PUSH RAFTEC", "Трійник натяжний ф25х16х25" (без редукційний!),
"Муфта натяжна ф20х16", "Гільза натяжна ф16" (НЕ насувна!). Гільзи обов'язкові!

КАРАКУЛІ МАЙСТРІВ: ⊥/Т=ТРІЙНИК; ∠/L=КОЛІНО; -//-=ПОВТОР типу; рядок без типу=продовження.
ПРОЕКТНІ ПОЗНАЧЕННЯ: Ду15/Dn15=1/2|Ду20=3/4|Ду25=1|Ду32=1¼|Ду50=2; ЗР=МРЗ; ВР=МРВ.
Групи PDF: "Труба Stabi Plus: Ekoplastik"+"32х4,4 м 40" → Труба PPR EVO ф32 Ekoplastik 40м.

КОМПЛЕКТИ ТП В2 (дефолт): SUR03+колектор STEEL+євроконуси+насос Termojet+крани BLACK+МРЗ ф25х3/4.
КОТЕЛ: крани кут.нак.гайкою ВВ+ВЗ (1/2 і 3/4)+фільтри BLACK+МРЗ ф20х1/2 х2+МРЗ ф25х3/4 х2.
РАДІАТОР бокове: клапан рад.кут.х2+компл.термостат.+МРЗ ф20х1/2+Hidros тип22.
РАДІАТОР VK: Hidros VK+вент.вставка+термоголовка WHITE+вузол нижн.підкл.1/2х3/4+муфта євроконус ф20х3/4.
"""


def normalize_photo(image_b64: str, caption: str = "", client_prefs: dict = None) -> list[dict]:
    rules = get_rules()
    rules_block = f"\nПравила менеджера:\n{rules}" if rules else ""
    brand_map = parse_caption_brands(caption)

    brand_hint = ""
    if brand_map:
        lines = [f"  {cat} → {toks[0]}" for cat, toks in brand_map.items()]
        brand_hint = "\n\n⚠️ ВИРОБНИКИ (суворо!):\n" + "\n".join(lines)
        brand_hint += "\nПриклад: каналізація→ostendorf значить ВСЯ каналізація OSTENDORF (НЕ ASG!)"

    ocr_block = get_ocr_prompt_block()
    prompt = f"""Ти — експерт сантехніки України. Рукописний список замовлення.
ПІДКАЗКА: {caption}{brand_hint}{rules_block}{ocr_block}
ЗНАННЯ: {ЗНАННЯ}

ЗАВДАННЯ: прочитай кожен рядок, нормалізуй назву (КОРОТКО!), витягни кількість.
JSON масив ТІЛЬКИ:
[{{"original":"що написано","normalized":"коротка назва","qty":"кількість",
"category":"plastic_ppr/sewage/push_systems/shutoff_valves/pumps/radiators_radiatorsvalve/filtration/insulation/metal_plastic/adapters_reducers/other",
"type":"труба/коліно/трійник/муфта/кран/гільза/перехід/...","dia":[110,50],"angle":87,"thread":"1/2 або null"}}]
type=тип виробу ОДНИМ словом; dia=ВСІ діаметри числами; angle=кут або null; thread=різьба або null."""

    try:
        image_bytes = base64.b64decode(image_b64)
        resp = _gemini_call([
            genai_types.Part.from_bytes(data=image_bytes, mime_type="image/jpeg"),
            genai_types.Part.from_text(text=prompt)])
        raw = resp.text.strip().replace('```json','').replace('```','').strip()
        if '[' in raw and ']' in raw:
            raw = raw[raw.index('['):raw.rindex(']')+1]
        return json.loads(raw)
    except Exception as e:
        return [{"original": f"Помилка OCR: {e}", "normalized": "", "qty": ""}]

def normalize_text(text: str, caption: str = "") -> list[dict]:
    """Нормалізація текстового запиту (команда 'пошук')."""
    rules = get_rules()
    rules_block = f"\nПравила менеджера:\n{rules}" if rules else ""
    ocr_block = get_ocr_prompt_block()
    prompt = f"""Ти — експерт сантехніки України. Текстовий запит менеджера.
ПІДКАЗКА: {caption}{rules_block}{ocr_block}
ЗНАННЯ: {ЗНАННЯ}

ЗАПИТ: {text}

Розбий на позиції, нормалізуй (КОРОТКО!), витягни кількість.
JSON масив ТІЛЬКИ:
[{{"original":"...","normalized":"...","qty":"...","category":"...",
"type":"тип одним словом","dia":[25],"angle":null,"thread":"3/4 або null"}}]"""
    try:
        resp = _gemini_call([genai_types.Part.from_text(text=prompt)])
        raw = resp.text.strip().replace('```json','').replace('```','').strip()
        if '[' in raw and ']' in raw:
            raw = raw[raw.index('['):raw.rindex(']')+1]
        return json.loads(raw)
    except Exception as e:
        return [{"original": text, "normalized": text, "qty": "", "category": "other"}]


def normalize_pdf(pdf_b64: str, caption: str = "") -> list[dict]:
    """Специфікація з PDF (проектна документація) — Gemini читає PDF нативно."""
    rules = get_rules()
    rules_block = f"\nПравила менеджера:\n{rules}" if rules else ""
    brand_map = parse_caption_brands(caption)
    brand_hint = ""
    if brand_map:
        lines = [f"  {cat} → {toks[0]}" for cat, toks in brand_map.items()]
        brand_hint = "\n⚠️ ВИРОБНИКИ (суворо!):\n" + "\n".join(lines)

    prompt = f"""Ти — експерт сантехніки. Це ПРОЕКТНА СПЕЦИФІКАЦІЯ (PDF, розділ ОВ).
ПІДКАЗКА: {caption}{brand_hint}{rules_block}
ЗНАННЯ: {ЗНАННЯ}

ЗАВДАННЯ: знайди таблиці специфікації (Найменування | Тип | Виробник | Од | Кількість).
Витягни КОЖНУ позицію. Пам'ятай:
- Групи ("Труба Stabi Plus: Ekoplastik" + підрядки розмірів) → успадковуй тип+виробника
- Виробник з колонки → у normalized
- section = розділ/блок позиції ("До П1", "Арматура", "Опалення", "Фітинги"...)
- Ду→дюйми, ЗР→МРЗ, ВР→МРВ
- Вентиляційне (Vents, повітроводи) включай теж — не знайдеться, це нормально

JSON масив ТІЛЬКИ:
[{{"original":"як у специфікації","normalized":"коротка назва","qty":"к-ть з од","category":"...","section":"розділ",
"type":"тип одним словом","dia":[32],"angle":null,"thread":null}}]"""
    try:
        pdf_bytes = base64.b64decode(pdf_b64)
        resp = _gemini_call([
            genai_types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf"),
            genai_types.Part.from_text(text=prompt)])
        raw = resp.text.strip().replace('```json','').replace('```','').strip()
        if '[' in raw and ']' in raw:
            raw = raw[raw.index('['):raw.rindex(']')+1]
        return json.loads(raw)
    except Exception as e:
        return [{"original": f"Помилка PDF: {e}", "normalized": "", "qty": "", "category": "other"}]


# ═══════════════════════════════════════════════════════════════════════════════
# CLAUDE ВИБІР
# ═══════════════════════════════════════════════════════════════════════════════
def _parse_claude_json(raw: str) -> list:
    raw = re.sub(r'```\w*', '', raw).strip()
    start, end = raw.find('['), raw.rfind(']') + 1
    if start != -1 and end > 0:
        try: return json.loads(raw[start:end])
        except Exception: pass
    objects = []
    for m in re.finditer(r'\{[^{}]*\}', raw[start:] if start != -1 else raw):
        try: objects.append(json.loads(m.group()))
        except Exception: pass
    return objects

def claude_pick_batch(позиції: list[dict], _retry=True) -> list[dict]:
    запити = []
    for i, пос in enumerate(позиції):
        brand_note = f"\n   ⚠️ ТІЛЬКИ: {пос['required_brand']}" if пос.get('required_brand') else ""
        кандидати = "\n".join(f"  {j+1}. [{c.get('_match_pct',0)}%] {c['name']}"
                               for j, c in enumerate(пос['candidates']))
        запити.append(f"{i+1}. {пос['normalized']}{brand_note}\n{кандидати}")

    prompt = f"""Сантехнік. Для кожного запиту — номер кандидата.
{chr(10).join(запити)}
Правила: діаметр ОБОВ'ЯЗКОВО збігається; ВИРОБНИК якщо вказано — тільки він; (п/з) тільки як останній варіант.
JSON рівно {len(позиції)} елементів:
[{{"знайдено":true,"номер_кандидата":1,"confidence":95,"reason":"причина","fail_reason":""}}]"""

    try:
        resp = claude.messages.create(
            model="claude-sonnet-4-5", max_tokens=4000,
            messages=[{"role": "user", "content": prompt}]
        )
        parsed = _parse_claude_json(resp.content[0].text)
        if not parsed: raise ValueError("порожній JSON")
        while len(parsed) < len(позиції):
            parsed.append({"знайдено": False, "confidence": 0, "reason": "", "fail_reason": "немає відповіді"})
        return parsed[:len(позиції)]
    except Exception as e:
        if _retry and len(позиції) > 1:
            print(f"⚠️ Батч впав ({e}), ретрай поштучно...")
            results = []
            for п in позиції:
                results.extend(claude_pick_batch([п], _retry=False))
            return results
        return [{"знайдено": False, "confidence": 0, "reason": "", "fail_reason": f"Claude: {e}"}] * len(позиції)

# ═══════════════════════════════════════════════════════════════════════════════
# FIND ITEMS
# ═══════════════════════════════════════════════════════════════════════════════
def find_items(позиції: list[dict], progress_cb=None) -> list[dict]:
    """
    4 РІВНІ ПРІОРИТЕТУ:
      1. Слова менеджера (підказка) — перебиває ВСЕ, включно з кешами
      1.5 Виробник у самому рядку майстра (WILO, Bonomi, Herz...)
      2. Кеш клієнта (тільки якщо не суперечить рівню 1)
      3. Кеш бота (тільки якщо не суперечить рівню 1)
      4. Профіль клієнта → дефолти → вільний пошук
    Плюс: бан-фільтр, шрабер-фільтр, retry-аналог з ⚠️.
    """
    результати = [None] * len(позиції)
    потребують_claude = []
    retry_позиції = []

    for i, пос in enumerate(позиції):
        if progress_cb: progress_cb(i+1, len(позиції))
        original     = пос.get('original', '')
        normalized   = пос.get('normalized', '')
        category     = пос.get('category', 'other')
        brand_map    = пос.get('_brand_map', {})
        client_slug  = пос.get('_client_slug')
        client_prefs = пос.get('_client_prefs', {})
        manager_brand = brand_map.get(category)
        # Якщо Gemini дав неточну категорію — пробуємо суміжні
        if not manager_brand and brand_map:
            for similar_cat in SIMILAR_CATS.get(category, []):
                if similar_cat in brand_map:
                    manager_brand = brand_map[similar_cat]
                    break

        # РІВЕНЬ 1.5: виробник у самому рядку (шукаємо ТІЛЬКИ в original)
        line_brand = None
        _orig_lc = original.lower()
        for bk, bt in BRAND_TOKENS.items():
            if re.search(r'(?<![a-zа-яёіїєґ0-9])' + re.escape(bk) + r'(?![a-zа-яёіїєґ0-9])', _orig_lc):
                line_brand = bt
                break
        hard_brand = manager_brand or line_brand   # жорсткий виробник рівня 1/1.5

        # РІВЕНЬ 2: кеш клієнта (поважає hard_brand)
        if client_slug:
            c = clients.client_cache_lookup(client_slug, original,
                                            required_brand_tokens=hard_brand)
            # Бан адміна (бот-кеш) перекриває клієнтський кеш
            if c and cache_is_banned(original, c.get('catalog_name','')):
                c = None
            # Валідація кешу: діаметр/тип мусять збігатись із запитом
            if c:
                _qa_c = пос.get('_qa') or build_qa(пос)
                пос['_qa'] = _qa_c
                if not validate_pick(_qa_c, {'name': c['catalog_name']}):
                    c = None
            if c:
                результати[i] = {
                    'original': original, 'normalized': normalized,
                    'знайдено': True, 'назва': c['catalog_name'],
                    'назва_повна': '', 'артикул': '', 'ціна': '',
                    'qty': пос.get('qty',''), 'category': c.get('category', category),
                    'confidence': c.get('confidence', 0), 'keyword_pct': 100,
                    'джерело': '👤 кеш клієнта' + (' ✅' if c.get('status')=='confirmed' else ''),
                    'reason': f"З кешу клієнта ({c.get('confidence',0)}%)",
                    'fail_reason': '', 'candidates_debug': [], '_from_cache': True,
                }
                continue

        # РІВЕНЬ 3: кеш бота (banned вирізає cache_lookup; перевіряємо hard_brand)
        cached = cache_lookup(original, brand_map)
        if cached and cache_is_banned(original, cached.get('catalog_name','')):
            cached = None   # страховка: бан головніший за будь-який запис
        if cached:
            _qa_b = пос.get('_qa') or build_qa(пос)
            пос['_qa'] = _qa_b
            if not validate_pick(_qa_b, {'name': cached['catalog_name']}):
                cached = None   # кеш суперечить діаметру/типу запиту — у пошук
        if cached:
            ok = True
            if hard_brand:
                nl = cached.get('catalog_name','').lower()
                ok = any(t.lower() in nl for t in hard_brand)
            if ok:
                результати[i] = {
                    'original': original, 'normalized': cached.get('normalized', normalized),
                    'знайдено': True, 'назва': cached['catalog_name'],
                    'назва_повна': '', 'артикул': '', 'ціна': '',
                    'qty': пос.get('qty',''), 'category': cached.get('category', category),
                    'confidence': cached.get('confidence',0), 'keyword_pct': 100,
                    'джерело': '🤖 кеш бота' + (' ✅' if cached.get('status')=='confirmed' else ''),
                    'reason': f"З кешу ({cached.get('confidence',0)}%)",
                    'fail_reason': '', 'candidates_debug': [], '_from_cache': True,
                }
                continue
            # кеш суперечить підказці — ігноруємо, шукаємо заново

        # ── Пошук: фільтр виробника ВСЕРЕДИНІ keyword_search ──────────────────
        кандидати = []
        required_brand = None
        джерело = ''
        brand_warning = ''

        if hard_brand:
            кандидати = smart_search(пос, top_n=12, brand_tokens=hard_brand)
            if кандидати:
                required_brand = hard_brand[0]
                джерело = '👨 менеджер' if manager_brand else '📝 з рядка'
            else:
                кандидати = smart_search(пос, top_n=12)
                brand_warning = f"⚠️ {hard_brand[0]} відсутній для цієї позиції"
                джерело = '⚠️ fallback'
        else:
            # РІВЕНЬ 4а: преференції клієнта з історії
            for brand, _cnt in client_prefs.get('by_category', {}).get(category, [])[:3]:
                bt = BRAND_TOKENS.get(brand)
                if not bt: continue
                кандидати = smart_search(пос, top_n=12, brand_tokens=bt)
                if кандидати:
                    required_brand = bt[0]
                    джерело = '👤 профіль клієнта'
                    break
            # РІВЕНЬ 4б: дефолти
            if not кандидати:
                for pt in DEFAULT_BRAND_PRIORITY.get(category, []):
                    кандидати = smart_search(пос, top_n=12, brand_tokens=pt)
                    if кандидати:
                        required_brand = pt[0]
                        джерело = '⚙️ дефолт'
                        break
            if not кандидати:
                кандидати = smart_search(пос, top_n=12)
                джерело = '🔍 вільний'

        if not кандидати:
            результати[i] = {**пос, 'знайдено': False, 'назва': '', 'артикул': '',
                              'ціна': '', 'confidence': 0, 'джерело': '',
                              'reason': '', 'fail_reason': 'не знайдено кандидатів',
                              'candidates_debug': []}
            continue

        # Бан-фільтр (позначені адміном як помилка)
        кандидати = [c for c in кандидати if not cache_is_banned(original, c['name'])]
        # Шрабер = інструмент, не фітинг
        if re.search(r'муфт|перех|редукц', normalized.lower()):
            кандидати = [c for c in кандидати if 'шрабер' not in c['name'].lower()]
        if not кандидати:
            результати[i] = {**пос, 'знайдено': False, 'назва': '', 'артикул': '',
                              'ціна': '', 'confidence': 0, 'джерело': '',
                              'reason': '', 'fail_reason': 'всі кандидати забанені',
                              'candidates_debug': []}
            continue

        # ⚡ АВТО-ПРИЙОМ: очевидний збіг → без Claude (швидше, дешевше, надійніше)
        # Умови: топ-1 ≥95%, ВСІ числа запиту є в назві, відрив від №2 ≥25%,
        # і виробник не в fallback (інакше потрібне попередження Claude-шляху).
        if кандидати and not brand_warning:
            top = кандидати[0]
            q_toks = tokenize(normalized)
            q_nums = {t for t in q_toks if t[0].isdigit()}
            top_toks = top.get('_tokens', set())
            gap_ok = (len(кандидати) == 1 or
                      top.get('_match_pct', 0) - кандидати[1].get('_match_pct', 0) >= 25)
            _qa_auto = пос.get('_qa') or build_qa(пос)
            attr_perfect = (top.get('_attr_tier') == 1 and validate_pick(_qa_auto, top))
            if attr_perfect or (q_nums and q_nums.issubset(top_toks)
                    and top.get('_match_pct', 0) >= 95 and gap_ok
                    and validate_pick(_qa_auto, top)):
                результати[i] = {
                    'original': original, 'normalized': normalized,
                    'знайдено': True, 'назва': top['name'],
                    'назва_повна': top.get('name_full', top['name']),
                    'артикул': top.get('artikul', ''),
                    'ціна': top.get('price', ''), 'qty': пос.get('qty', ''),
                    'category': category, 'confidence': 95,
                    'keyword_pct': top.get('_match_pct', 0),
                    'джерело': '⚡ точний збіг', 'brand_warning': '',
                    'reason': 'Всі розміри і назва збіглись — вибрано без AI',
                    'fail_reason': '',
                    'candidates_debug': [c['name'] for c in кандидати[:3]],
                }
                cache_save(original, brand_map, normalized, top['name'], category, 95)
                if client_slug:
                    clients.client_cache_save(client_slug, original, top['name'], category, 95)
                continue

        потребують_claude.append({
            'idx': i, 'normalized': normalized, 'original': original,
            'candidates': кандидати,
            'candidates_debug': [c['name'] for c in кандидати[:5]],
            'qty': пос.get('qty',''), 'required_brand': required_brand,
            'category': category, 'brand_map': brand_map,
            'client_slug': client_slug, 'джерело': джерело,
            'brand_warning': brand_warning,
        })

    if потребують_claude:
        відповіді = claude_pick_batch(потребують_claude)
        for j, пос in enumerate(потребують_claude):
            r = відповіді[j] if j < len(відповіді) else {'знайдено': False}
            idx = пос['idx']
            conf = int(r.get('confidence', 0))
            if r.get('знайдено') and r.get('номер_кандидата'):
                n = max(0, min(int(r['номер_кандидата'])-1, len(пос['candidates'])-1))
                found = пос['candidates'][n]
                # ПОСТ-ВАЛІДАЦІЯ: Claude міг впевнено помилитись — код перевіряє
                _qa = пос.get('_qa') or build_qa(пос)
                if not validate_pick(_qa, found):
                    _alt = next((c for c in пос['candidates'] if validate_pick(_qa, c)), None)
                    if _alt is not None:
                        found = _alt
                        r['confidence'] = min(int(r.get('confidence', 0)), 75)
                        r['reason'] = (r.get('reason','') + ' | ⚙️ авто-заміна: валідація діаметр/тип')[:120]
                    else:
                        r['знайдено'] = False
                        r['fail_reason'] = 'валідація: діаметр/тип запиту відсутній у всіх кандидатах'
            if r.get('знайдено') and r.get('номер_кандидата'):
                reason = r.get('reason', '')
                if пос.get('brand_warning'):
                    reason = f"{пос['brand_warning']}. {reason}"
                результати[idx] = {
                    'original': пос['original'], 'normalized': пос['normalized'],
                    'знайдено': True, 'назва': found['name'],
                    'назва_повна': found.get('name_full', found['name']),
                    'артикул': found.get('artikul', ''),
                    'ціна': found.get('price', ''), 'qty': пос['qty'],
                    'category': пос['category'], 'confidence': conf,
                    'keyword_pct': found.get('_match_pct', 0),
                    'джерело': пос['джерело'],
                    'brand_warning': пос.get('brand_warning',''),
                    'reason': reason, 'fail_reason': '',
                    'candidates_debug': пос['candidates_debug'],
                }
                # Кешуємо успіх (auto, confirmed/banned не перезаписуються)
                cache_save(пос['original'], пос['brand_map'], пос['normalized'],
                           found['name'], пос['category'], conf)
                if пос.get('client_slug'):
                    clients.client_cache_save(пос['client_slug'], пос['original'],
                                              found['name'], пос['category'], conf)
            else:
                результати[idx] = {
                    'original': пос['original'], 'normalized': пос['normalized'],
                    'знайдено': False, 'назва': '', 'артикул': '', 'ціна': '',
                    'qty': пос['qty'], 'category': пос['category'],
                    'confidence': conf, 'джерело': '',
                    'reason': '', 'fail_reason': r.get('fail_reason', 'не знайдено'),
                    'candidates_debug': пос['candidates_debug'],
                }
                # Кандидат на другий шанс (аналог іншого виробника)
                if пос.get('required_brand'):
                    retry_позиції.append(пос)

        # ═══ ДРУГИЙ ШАНС: у виробника немає → аналог з явним ⚠️ ═══
        if retry_позиції:
            retry_batch = []
            for пос in retry_позиції:
                nc = smart_search(пос, top_n=12)
                nc = [c for c in nc if not cache_is_banned(пос['original'], c['name'])]
                if re.search(r'муфт|перех|редукц', пос['normalized'].lower()):
                    nc = [c for c in nc if 'шрабер' not in c['name'].lower()]
                _rb = пос['required_brand'].lower()
                nc2 = [c for c in nc if _rb not in c['name'].lower()]
                nc = nc2 or nc
                if nc:
                    retry_batch.append({**пос, 'candidates': nc,
                        'candidates_debug': [c['name'] for c in nc[:5]],
                        'required_brand': None, 'old_brand': пос['required_brand']})
            if retry_batch:
                відп2 = claude_pick_batch(retry_batch)
                for j, пос in enumerate(retry_batch):
                    r = відп2[j] if j < len(відп2) else {'знайдено': False}
                    if r.get('знайдено') and r.get('номер_кандидата'):
                        n = max(0, min(int(r['номер_кандидата'])-1, len(пос['candidates'])-1))
                        found = пос['candidates'][n]
                        warn = f"⚠️ у {пос['old_brand']} немає — аналог"
                        результати[пос['idx']] = {
                            'original': пос['original'], 'normalized': пос['normalized'],
                            'знайдено': True, 'назва': found['name'],
                            'назва_повна': found.get('name_full', found['name']),
                            'артикул': found.get('artikul',''),
                            'ціна': found.get('price',''), 'qty': пос['qty'],
                            'category': пос['category'],
                            'confidence': int(r.get('confidence',0)),
                            'keyword_pct': found.get('_match_pct',0),
                            'джерело': '⚠️ аналог', 'brand_warning': warn,
                            'reason': f"{warn}. {r.get('reason','')}",
                            'fail_reason': '',
                            'candidates_debug': пос['candidates_debug'],
                        }
                        # аналоги НЕ кешуємо

    # Ціна/артикул/повна назва для кешованих результатів
    for r in результати:
        if r and r.get('_from_cache') and r.get('назва'):
            for it in CATALOG:
                if it['name'] == r['назва']:
                    r['ціна'] = it.get('price','')
                    r['артикул'] = it.get('artikul','')
                    r['назва_повна'] = it.get('name_full', it['name'])
                    break
            r.pop('_from_cache', None)
    return результати

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
def parse_qty(s):
    s = str(s or '').strip()
    if not s: return '', ''
    m = re.match(r'(\d+(?:[.,]\d+)?)\s*(.*)', s)
    if m:
        try:
            num = float(m.group(1).replace(',', '.'))
            num = int(num) if num == int(num) else num
        except Exception:
            num = m.group(1)
        return num, m.group(2).strip().rstrip('.').strip()
    return s, ''

def create_excel(результати: list[dict]):
    from openpyxl.styles import PatternFill
    RED    = PatternFill('solid', fgColor='FFC7CE')
    YELLOW = PatternFill('solid', fgColor='FFF3B0')

    rows, flags = [], []
    not_found, warn = [], []

    for r in результати:
        if not r: continue
        conf, kw = r.get('confidence', 0), r.get('keyword_pct', 0)
        qty_num, qty_unit = parse_qty(r.get('qty', ''))

        if r.get('знайдено'):
            suspicious = conf < 70 or kw < 50 or r.get('brand_warning') or r.get('джерело','') in ('⚠️ fallback','🔍 вільний','⚠️ аналог')
            rows.append({
                '№':            len(rows) + 1,
                'Артикул':      r.get('артикул', ''),
                'Наименование': r.get('назва_повна') or r.get('назва', ''),
                'Кількість':    qty_num, 'Од.': qty_unit,
                'Ціна':         r.get('ціна', ''),
                'Збіг':         f"🔍{kw}%/🤖{conf}%",
                'Джерело':      r.get('джерело', ''),
                'Розділ':       r.get('розділ', ''),
                'Чому знайшло': r.get('reason', ''),
                'Оригінал':     r.get('original', ''),
            })
            flags.append('warn' if suspicious else '')
            if suspicious: warn.append(r.get('original',''))
        else:
            cands = r.get('candidates_debug', [])
            best = cands[0] if cands else ''
            art, full, price = '', best, ''
            for it in CATALOG:
                if it['name'] == best:
                    art, full, price = it.get('artikul',''), it.get('name_full', best), it.get('price','')
                    break
            rows.append({
                '№': len(rows) + 1,
                'Артикул': art, 'Наименование': full,
                'Кількість': qty_num, 'Од.': qty_unit,
                'Ціна': price, 'Збіг': '—',
                'Джерело': '❓ НЕ ЗНАЙДЕНО',
                'Розділ': r.get('розділ', ''),
                'Чому знайшло': (r.get('fail_reason','') or '')[:100],
                'Оригінал': r.get('original', ''),
            })
            flags.append('nf')
            not_found.append(r.get('original', ''))

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df = pd.DataFrame(rows) if rows else pd.DataFrame(
            columns=['№','Артикул','Наименование','Кількість','Од.','Ціна','Збіг','Джерело','Розділ','Чому знайшло','Оригінал'])
        df.to_excel(writer, index=False, sheet_name='Замовлення')
        ws = writer.sheets['Замовлення']
        # Ширина колонок: № вузька, Наименование широка
        ws.column_dimensions['A'].width = 4   # №
        ws.column_dimensions['C'].width = 55  # Наименование
        ws.column_dimensions['J'].width = 20  # Оригінал
        for i, fl in enumerate(flags, start=2):
            fill = RED if fl == 'nf' else (YELLOW if fl == 'warn' else None)
            if fill:
                for cell in ws[i]: cell.fill = fill
    output.seek(0)
    return output, not_found, warn

# ═══════════════════════════════════════════════════════════════════════════════
# СТАНИ (визначаємо ДО хендлерів і process_batch!)
# ═══════════════════════════════════════════════════════════════════════════════
user_batches  = {}
stop_flags    = {}
pending_hints = {}
last_results  = {}
_fix_state    = {}
_train_state  = {}
_text_pending = {}   # chat_id → текст що чекає рішення список/підказка
_manual_wait  = {}   # chat_id → {'mode':'fix'|'train'} чекаємо ручну назву
BATCH_TIMEOUT = 4

def safe_edit(chat_id, msg_id, text):
    try: bot.edit_message_text(text, chat_id=chat_id, message_id=msg_id)
    except Exception: pass

# ═══════════════════════════════════════════════════════════════════════════════
# PROCESS BATCH
# ═══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
# АВТО-РОЗГОРТАННЯ: маркер PUSH, "+ізол" → утеплювач, гільзи до PUSH-фітингів
# ═══════════════════════════════════════════════════════════════════════════════

# Труба → утеплювач (від Тараса): 16→18, 20→22, 25→28, 32→35
INSUL_DIA_MAP = {16: 18, 20: 22, 25: 28, 32: 35, 40: 42}

def _qty_num(qty_str):
    m = re.search(r'(\d+(?:[.,]\d+)?)', str(qty_str or ''))
    return float(m.group(1).replace(',', '.')) if m else 0

def expand_push_marker(позиції):
    """Слово 'гільз' будь-де в списку = все замовлення PUSH (гільзи лише там)."""
    has_sleeve = any('гільз' in (п.get('original','') + п.get('normalized','')).lower()
                     for п in позиції)
    if not has_sleeve:
        return позиції
    for п in позиції:
        if п.get('category') in ('plastic_ppr', 'metal_plastic', 'other'):
            п['category'] = 'push_systems'
            n = п.get('normalized', '')
            if 'push' not in n.lower() and 'натяжн' not in n.lower():
                п['normalized'] = (n + ' натяжний PUSH').strip()
    return позиції

def expand_insulation(позиції):
    """'труба 25 - 40м + ізол' → труба 40м + утеплювач PLM синій 20м + червоний 20м."""
    out = []
    for п in позиції:
        out.append(п)
        orig = п.get('original', '').lower()
        qa = п.get('_qa') or build_qa(п)
        п['_qa'] = qa
        if qa.get('type') == 'труба' and re.search(r'ізол|изол|утепл', orig):
            dia = (qa.get('dia') or [None])[0]
            ins_dia = INSUL_DIA_MAP.get(dia)
            m_total = _qty_num(п.get('qty'))
            if ins_dia and m_total > 0:
                half = m_total / 2
                half_s = str(int(half)) if half == int(half) else f"{half:.1f}"
                for color in ('синій', 'червоний'):
                    out.append({
                        'original': f"(авто +ізол) утеплювач ф{ins_dia} {color}",
                        'normalized': f"Утеплювач ламін. для труб ф {ins_dia}х6 {color} PLM",
                        'qty': f"{half_s} м", 'category': 'insulation',
                        'type': 'утеплювач', 'dia': [ins_dia],
                        'section': п.get('section', ''),
                    })
    return out

# Трубні виходи фітинга (різьбовий вихід гільзи НЕ потребує)
def _push_outlets(п):
    qa = п.get('_qa') or build_qa(п)
    п['_qa'] = qa
    typ = qa.get('type')
    text = f"{п.get('normalized','')} {п.get('original','')}"
    # Всі числа в групі розмірів NхMхK (порядок і повтори важливі!)
    g = re.search(r'(\d{2})\s*[хx×]\s*(\d{2})(?:\s*[хx×]\s*(\d{2}))?', text)
    dims = [int(x) for x in g.groups() if x] if g else list(qa.get('dia') or [])
    has_thread = bool(qa.get('thread')) or bool(re.search(r'мрз|мрв|рз|вр|різьб', text.lower()))
    if typ == 'трійник':
        outs = dims if len(dims) == 3 else (dims * 3)[:3] if dims else []
    elif typ in ('коліно',):
        outs = dims if len(dims) == 2 else (dims * 2)[:2] if dims else []
    elif typ in ('муфта', 'перехід'):
        if has_thread:
            outs = dims[:1]                     # МРЗ 25х3/4 → 1 гільза ф25
        else:
            outs = dims if len(dims) == 2 else (dims * 2)[:2] if dims else []
    elif typ == 'заглушка':
        outs = dims[:1]
    else:
        outs = []
    return outs

def expand_push_sleeves(позиції):
    """Кожен PUSH-фітинг → гільза на кожен трубний вихід × кількість фітингів.
    Якщо майстер САМ написав гільзи — не дублюємо."""
    if any('гільз' in (п.get('original','') + п.get('normalized','')).lower()
           and (п.get('_qa') or build_qa(п)).get('type') == 'гільза'
           for п in позиції):
        return позиції
    sleeves = {}
    for п in позиції:
        if п.get('category') != 'push_systems':
            continue
        outs = _push_outlets(п)
        if not outs:
            continue
        n_fit = int(_qty_num(п.get('qty')) or 1)
        for d in outs:
            sleeves[d] = sleeves.get(d, 0) + n_fit
    for d in sorted(sleeves):
        позиції.append({
            'original': f"(авто) гільзи ф{d} до PUSH-фітингів",
            'normalized': f"Гільза натяжна ф {d} PUSH",
            'qty': f"{sleeves[d]} шт", 'category': 'push_systems',
            'type': 'гільза', 'dia': [d],
        })
    return позиції


def process_batch(chat_id: int):
    batch = user_batches.pop(chat_id, None)
    if not batch: return
    stop_flags.pop(chat_id, None)
    items = batch['items']

    active_slug = clients.get_active(chat_id)
    client_prefs = clients.get_preferences(active_slug) if active_slug else {}
    client_name = ''
    if active_slug:
        _p = clients.get_profile(active_slug)
        client_name = _p['name'] if _p else active_slug
    client_line = f"\n👤 Клієнт: {client_name}" if client_name else ""

    status = bot.send_message(chat_id, f"⏳ Обробляю {len(items)} файл(ів)...{client_line}")
    msg_id = status.message_id

    всі_позиції, errors = [], []

    for idx, item in enumerate(items, 1):
        if stop_flags.get(chat_id):
            safe_edit(chat_id, msg_id, "🛑 Зупинено.")
            return
        try:
            safe_edit(chat_id, msg_id, f"📖 Читаю файл {idx}/{len(items)}...")
            if item['type'] == 'photo':
                позиції = normalize_photo(item['data'], item.get('caption',''))
            elif item['type'] == 'pdf':
                позиції = normalize_pdf(item['data'], item.get('caption',''))
            elif item['type'] == 'text':
                позиції = normalize_text(item['text'], item.get('caption',''))
            else:
                позиції = []
            всі_позиції.extend(позиції)
        except Exception as e:
            errors.append(f"❌ Файл {idx}: {e}")

    if not всі_позиції:
        safe_edit(chat_id, msg_id, "😕 Не розпізнано позицій.\n" + "\n".join(errors))
        return

    # Прикріплюємо brand_map
    # Авто-розгортання: PUSH-маркер → +ізол → гільзи
    всі_позиції = expand_push_marker(всі_позиції)
    всі_позиції = expand_insulation(всі_позиції)
    всі_позиції = expand_push_sleeves(всі_позиції)

    # Збираємо підказки з УСІХ елементів батчу (менеджер міг додати підказку до кожного фото)
    all_captions = [it.get('caption','') for it in items if it.get('caption','')]
    caption = ' | '.join(all_captions)  # об'єднуємо
    brand_map = parse_caption_brands(caption)
    for п in всі_позиції:
        п['_brand_map'] = brand_map
        if active_slug:
            п['_client_slug'] = active_slug
            п['_client_prefs'] = client_prefs

    safe_edit(chat_id, msg_id, f"🔍 Шукаю {len(всі_позиції)} позицій...")

    def progress(cur, total):
        if cur % 5 == 0 or cur == total:
            safe_edit(chat_id, msg_id, f"🔍 Пошук: {cur}/{total}...")

    результати = find_items(всі_позиції, progress_cb=progress)

    # Розділ (PDF-специфікації) → у результати
    for _r, _п in zip(результати, всі_позиції):
        if _r is not None:
            _r.setdefault('розділ', _п.get('section', ''))

    # Лог незнайдених для звіту "Діри каталогу"
    log_not_found([r for r in результати if r and not r.get('знайдено')])

    safe_edit(chat_id, msg_id, "📊 Формую Excel...")
    excel, not_found, warn = create_excel(результати)

    знайдено = [r for r in результати if r and r.get('знайдено')]
    total = len([r for r in результати if r])

    bot.send_document(chat_id, excel, visible_file_name="замовлення.xlsx")

    звіт = f"✅ Знайдено: {len(знайдено)}/{total}\n"
    if not_found: звіт += f"🟥 Не знайдено: {len(not_found)}\n"
    if warn: звіт += f"🟨 Перевір: {len(warn)}\n"
    if errors: звіт += "\n" + "\n".join(errors)
    safe_edit(chat_id, msg_id, звіт)

    # Кнопки навчання — для всіх (адмін застосовує одразу, інші → на розгляд)
    mk = InlineKeyboardMarkup()
    mk.add(InlineKeyboardButton("🎓 Навчання", callback_data="tr_go"),
           InlineKeyboardButton("✖️ Закрити",  callback_data="tr_close"))
    note = "" if chat_id == ADMIN_ID else "\n(твої виправлення підтверджує адмін)"
    bot.send_message(chat_id, f"Перевір файл. Якщо є помилки — тапни Навчання:{note}",
                     reply_markup=mk)

    last_results[chat_id] = {'результати': [r for r in результати if r], 'client_slug': active_slug}
    if active_slug:
        try:
            clients.save_order(active_slug, результати, caption)
        except Exception as e:
            print(f"⚠️ Історія клієнта: {e}")
    username = items[0].get('username', str(chat_id)) if items else str(chat_id)
    log_usage(chat_id, username, total, len(знайдено), len(items))
    try:
        storage.save_now()
    except Exception as _e:
        print(f"⚠️ storage: {_e}")

def add_to_batch(chat_id: int, item: dict):
    if chat_id not in user_batches:
        user_batches[chat_id] = {'items': []}
        bot.send_message(chat_id, "📥 Отримав! Чекаю ще файли (4 сек)...")
    if 'timer' in user_batches[chat_id]:
        user_batches[chat_id]['timer'].cancel()
    user_batches[chat_id]['items'].append(item)
    t = threading.Timer(BATCH_TIMEOUT, process_batch, args=[chat_id])
    user_batches[chat_id]['timer'] = t
    t.daemon = True
    t.start()

# ═══════════════════════════════════════════════════════════════════════════════
# КЛАВІАТУРА
# ═══════════════════════════════════════════════════════════════════════════════
def main_keyboard(uid: int) -> ReplyKeyboardMarkup:
    kb = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add(KeyboardButton("📸 Як користуватись"), KeyboardButton("🛑 Стоп"))
    kb.add(KeyboardButton("📋 Правило"), KeyboardButton("📊 Кеш"))
    kb.add(KeyboardButton("👥 Клієнти"), KeyboardButton("👥 Кеш клієнта"))
    if is_admin(uid):
        kb.add(KeyboardButton("👑 Статистика"), KeyboardButton("👑 Правила на розгляд"))
        kb.add(KeyboardButton("👑 Логи"), KeyboardButton("👑 Діри каталогу"))
    return kb

# ═══════════════════════════════════════════════════════════════════════════════
# ХЕНДЛЕРИ
# ═══════════════════════════════════════════════════════════════════════════════
@bot.message_handler(commands=['start', 'help'])
def handle_start(message):
    admin = is_admin(message.from_user.id)
    admin_note = ""
    if admin:
        admin_note = ("\n\n👑 *Адмін:* твої виправлення застосовуються одразу. "
                      "Чужі — чекають у «👑 Правила на розгляд».")
    else:
        admin_note = "\n\n_Твої виправлення і правила підтверджує адмін._"
    bot.reply_to(message, f"""👋 Привіт! Я підбираю сантехніку з бази по фото списку.

*📋 ЯК ПРАЦЮВАТИ (3 кроки):*
1️⃣ Напиши виробників (кожен рядок = категорія):
`каналізація остендорф`
`пайка екопластик`
`крани рафтек`
2️⃣ Кинь фото рукописного списку (можна кілька)
3️⃣ Отримай Excel: 🟥 не знайдено, 🟨 перевір

*👤 ПОСТІЙНІ КЛІЄНТИ (щоб бот пам'ятав звички):*
`новий клієнт Петренко` — створити профіль (один раз)
`клієнт Петренко` — активуй ПЕРЕД фото! Бот візьме
його минулі підбори і улюблених виробників.
`клієнт стоп` — вимкнути | `клієнти` — список

*🎓 ЯКЩО БОТ ПОМИЛИВСЯ:*
Після файлу тапни «🎓 Навчання» → напиши номери
неправильних рядків (напр: `3 7 12`) → вибери що не так
→ тапни правильний варіант. Все!
Або швидко: `вірно 3` / `помилка 5` / `виправ 5`{admin_note}

`пошук <текст>` — підбір без фото | /stop — зупинити""",
        parse_mode="Markdown",
        reply_markup=main_keyboard(message.from_user.id))

@bot.message_handler(func=lambda m: m.text == "📸 Як користуватись")
def kb_howto(message):
    bot.reply_to(message, """📸 *ПОВНА ІНСТРУКЦІЯ*

*Крок 1. Клієнт (якщо постійний):*
`клієнт Петренко` — і бот згадає ВСІ його минулі
замовлення: які виробники брав, які саме товари.
Повторний список знайдеться швидше і точніше.
Новий? → `новий клієнт Петренко`

*Крок 2. Виробники:*
Напиши повідомлення-підказку, кожен рядок окремо:
`каналізація остендорф`
`пайка екопластик`
`пуш рафтек`
Або одним словом на все: `усе рафтек`

*Крок 3. Фото:*
Кинь фото списку (можна одразу кілька — почекай
4 сек, бот збере їх в одне замовлення).

*Крок 4. Перевір Excel:*
🟥 червоний = не знайдено (в рядку — найближчий аналог)
🟨 жовтий = сумнівно, глянь оком
Колонка «Джерело» показує звідки взявся вибір.

*Крок 5. Навчи якщо є помилки:*
Тапни «🎓 Навчання» під файлом → номери рядків
(`3 7 12`) → причину → правильний варіант.
Наступного разу бот вже не помилиться!""", parse_mode="Markdown")

@bot.message_handler(func=lambda m: m.text in ("🛑 Стоп", "🛑 стоп"))
def kb_stop(message):
    handle_stop(message)

@bot.message_handler(func=lambda m: m.text in ("📋 Правило",))
def kb_rule_btn(message):
    bot.reply_to(message, "Напиши: `правило <текст>`\nПриклад: `правило рожон = трійник`", parse_mode="Markdown")

@bot.message_handler(func=lambda m: m.text == "📊 Кеш")
def kb_cache(message):
    handle_cache_info(message)

@bot.message_handler(func=lambda m: m.text == "👥 Клієнти")
def kb_clients(message):
    handle_clients_list(message)

@bot.message_handler(func=lambda m: m.text == "👥 Кеш клієнта")
def kb_client_cache(message):
    slug = clients.get_active(message.chat.id)
    if not slug:
        bot.reply_to(message, "Немає активного клієнта.\nАктивуй: `клієнт <ім'я>`", parse_mode="Markdown")
        return
    p = clients.get_profile(slug)
    name = p['name'] if p else slug
    cache = clients.get_client_cache(slug)
    if not cache:
        bot.reply_to(message, f"👥 Кеш *{name}* порожній.", parse_mode="Markdown")
        return
    icons = {'confirmed':'✅','banned':'❌','auto':'🔹'}
    lines = []
    for k, v in list(cache.items())[-10:]:
        lines.append(f"{icons.get(v.get('status','auto'),'🔹')} `{k[:35]}` → {v.get('catalog_name','')[:45]}")
    bot.reply_to(message,
        f"👥 Кеш *{name}*: {len(cache)} записів\n✅ підтв | ❌ бан | 🔹 авто\n\n" + "\n".join(lines),
        parse_mode="Markdown")

@bot.message_handler(func=lambda m: m.text == "👑 Статистика")
def kb_stats(message):
    if not is_admin(message.from_user.id): return
    bot.reply_to(message, get_usage_stats(), parse_mode="Markdown")

@bot.message_handler(func=lambda m: m.text == "👑 Логи")
def kb_logs(message):
    if not is_admin(message.from_user.id): return
    if not os.path.exists(USAGE_LOG_FILE):
        bot.reply_to(message, "Лог порожній.")
        return
    with open(USAGE_LOG_FILE, "rb") as f:
        bot.send_document(message.chat.id, f, visible_file_name="usage_log.json")

@bot.message_handler(func=lambda m: m.text == "👑 Діри каталогу")
def kb_gaps(message):
    if not is_admin(message.from_user.id):
        return
    bot.reply_to(message, get_catalog_gaps())

@bot.message_handler(commands=['діри', 'gaps'])
def handle_gaps(message):
    if not is_admin(message.from_user.id):
        return
    bot.reply_to(message, get_catalog_gaps())

@bot.message_handler(func=lambda m: m.text == "👑 Правила на розгляд")
def kb_pending(message):
    if not is_admin(message.from_user.id): return
    show_pending_rules(message.chat.id)

def show_pending_rules(chat_id):
    rules = load_pending_rules()
    fixes = load_pending_fixes()
    if not rules and not fixes:
        bot.send_message(chat_id, "✅ Немає нічого на розгляд.")
        return
    for i, r in enumerate(rules):
        mk = InlineKeyboardMarkup()
        mk.add(InlineKeyboardButton("✅ Підтвердити", callback_data=f"approve_{i}"),
               InlineKeyboardButton("❌ Відхилити",   callback_data=f"reject_{i}"))
        bot.send_message(chat_id,
            f"📋 Правило #{i+1} від {r['username']} ({r['date']}):\n{r['rule']}",
            reply_markup=mk)
    for i, f in enumerate(fixes):
        mk = InlineKeyboardMarkup()
        mk.add(InlineKeyboardButton("✅ Застосувати", callback_data=f"fixok_{i}"),
               InlineKeyboardButton("❌ Відхилити",   callback_data=f"fixno_{i}"))
        bot.send_message(chat_id,
            f"🎓 Виправлення #{i+1} від {f['username']} ({f.get('date','')}):\n"
            f"«{f.get('original','')[:45]}»\n"
            f"❌ {(f.get('old_name') or '—')[:55]}\n"
            f"✅ {(f.get('new_name') or '(тільки заборонити старе)')[:55]}",
            reply_markup=mk)
@bot.callback_query_handler(func=lambda c: c.data.startswith(('approve_','reject_')))
def handle_rule_decision(call):
    if not is_admin(call.from_user.id):
        bot.answer_callback_query(call.id, "⛔ Тільки адмін"); return
    action, idx = call.data.split('_')
    idx = int(idx)
    rules = load_pending_rules()
    if idx >= len(rules):
        bot.answer_callback_query(call.id, "Вже оброблено"); return
    rule = rules.pop(idx)
    save_pending_rules(rules)
    if action == 'approve':
        add_rule(rule['rule'])
        bot.edit_message_text(f"✅ ПІДТВЕРДЖЕНО:\n`{rule['rule']}`",
            call.message.chat.id, call.message.message_id, parse_mode="Markdown")
        try: bot.send_message(rule['user_id'], f"✅ Твоє правило підтверджено:\n`{rule['rule']}`", parse_mode="Markdown")
        except Exception: pass
    else:
        bot.edit_message_text(f"❌ ВІДХИЛЕНО:\n`{rule['rule']}`",
            call.message.chat.id, call.message.message_id, parse_mode="Markdown")
    bot.answer_callback_query(call.id, "Готово")


@bot.callback_query_handler(func=lambda c: c.data.startswith(('fixok_', 'fixno_')))
def handle_fixq_decision(call):
    if not is_admin(call.from_user.id):
        bot.answer_callback_query(call.id, "⛔ Тільки адмін")
        return
    action, idx = call.data.split('_')
    idx = int(idx)
    fixes = load_pending_fixes()
    if idx >= len(fixes):
        bot.answer_callback_query(call.id, "Вже оброблено")
        return
    fix = fixes.pop(idx)
    save_pending_fixes(fixes)
    who = fix.get('username', '?')
    if action == 'fixok':
        apply_fix(fix)
        bot.edit_message_text(
            f"✅ ЗАСТОСОВАНО (від @{who}):\n«{fix.get('original','')[:40]}»\n"
            f"❌ {(fix.get('old_name') or '—')[:50]}\n✅ {(fix.get('new_name') or 'бан')[:50]}",
            call.message.chat.id, call.message.message_id)
        try:
            bot.send_message(fix['user_id'],
                f"✅ Твоє виправлення підтверджено адміном:\n"
                f"«{fix.get('original','')[:40]}» → {(fix.get('new_name') or 'заборонено')[:55]}")
        except Exception:
            pass
    else:
        bot.edit_message_text(
            f"❌ ВІДХИЛЕНО (від @{who}):\n«{fix.get('original','')[:40]}»",
            call.message.chat.id, call.message.message_id)
        try:
            bot.send_message(fix['user_id'],
                f"❌ Твоє виправлення відхилено адміном:\n«{fix.get('original','')[:40]}»")
        except Exception:
            pass
    bot.answer_callback_query(call.id, "Готово")


@bot.callback_query_handler(func=lambda c: c.data in ("tr_go","tr_close"))
def tr_start(call):
    if call.data == "tr_close":
        try: bot.edit_message_text("✖️ Закрито.", call.message.chat.id, call.message.message_id)
        except Exception: pass
        bot.answer_callback_query(call.id); return
    _train_state[call.message.chat.id] = {'stage': 'rows'}
    try:
        bot.edit_message_text(
            "✍️ Напиши номери НЕПРАВИЛЬНИХ рядків через пробіл (напр: 3 7 12):",
            call.message.chat.id, call.message.message_id)
    except Exception: pass
    bot.answer_callback_query(call.id)

@bot.message_handler(func=lambda m: m.text and m.chat.id in _train_state
                     and _train_state[m.chat.id].get('stage') == 'rows'
                     and re.fullmatch(r'[\d\s,]+', m.text.strip()))
def tr_rows(message):
    last = last_results.get(message.chat.id)
    total = len(last['результати']) if last else 0
    nums = sorted({int(x) for x in re.findall(r'\d+', message.text)
                   if 1 <= int(x) <= total})
    if not nums:
        bot.reply_to(message, f"⚠️ Валідних номерів немає (всього {total} рядків).")
        return
    _train_state[message.chat.id] = {'stage': 'classify', 'rows': nums, 'i': 0}
    _tr_show_row(message.chat.id)

def _tr_show_row(chat_id):
    st = _train_state.get(chat_id)
    last = last_results.get(chat_id)
    if not st or not last: return
    if st['i'] >= len(st['rows']):
        bot.send_message(chat_id, f"✅ Навчання завершено! Оброблено {len(st['rows'])} рядків.")
        _train_state.pop(chat_id, None); return
    row = st['rows'][st['i']]
    r = last['результати'][row-1]
    mk = InlineKeyboardMarkup(row_width=1)
    mk.add(InlineKeyboardButton("📖 Бот неправильно ПРОЧИТАВ рядок", callback_data="tro"),
           InlineKeyboardButton("🎯 Прочитав вірно, товар НЕ ТОЙ",   callback_data="trg"),
           InlineKeyboardButton("⏭ Пропустити",                     callback_data="trs"))
    m = bot.send_message(chat_id,
        f"🎓 Рядок {row} ({st['i']+1}/{len(st['rows'])})\n"
        f"Написано: {r.get('original','')[:50]}\n"
        f"Бот дав: {(r.get('назва','') or '❓ не знайдено')[:60]}\n\nЩо не так?",
        reply_markup=mk)
    st['msg_id'] = m.message_id

@bot.callback_query_handler(func=lambda c: c.data in ("tro","trg","trw","trb","trc","trs","trn","trm") or c.data.startswith("trp_"))
def tr_classify(call):
    chat_id = call.message.chat.id
    st = _train_state.get(chat_id)
    last = last_results.get(chat_id)
    if not st or not last:
        bot.answer_callback_query(call.id, "Сесія застаріла"); return
    admin = is_admin(call.from_user.id)
    row = st['rows'][st['i']]
    r = last['результати'][row-1]
    original = r.get('original','')
    old_name = r.get('назва','')
    cat = r.get('category','other')
    cslug = last.get('client_slug')
    uname = call.from_user.username or str(call.from_user.id)

    def advance():
        st['i'] += 1
        _tr_show_row(chat_id)

    if call.data == "trs":
        bot.answer_callback_query(call.id, "Пропущено"); advance(); return

    if call.data == "trm":
        _manual_wait[chat_id] = {'mode': 'train'}
        bot.answer_callback_query(call.id)
        bot.send_message(chat_id,
            "✍️ Напиши назву товару (як у прайсі, можна частину):\n"
            "напр: `коліно 110 45 остендорф`", parse_mode="Markdown")
        return

    if call.data == "trn":
        if old_name:
            if admin:
                cache_ban_pair(original, old_name, cat)
                if cslug:
                    clients.client_cache_set_status(cslug, original, old_name, 'banned')
                bot.answer_callback_query(call.id, "❌ Забанено")
            else:
                n = add_pending_fix({'original': original, 'old_name': old_name,
                                     'new_name': None, 'category': cat,
                                     'client_slug': cslug,
                                     'normalized': r.get('normalized',''),
                                     'user_id': call.from_user.id, 'username': uname})
                notify_admin_fix(uname, original, old_name, None, n)
                bot.answer_callback_query(call.id, "📥 Надіслано адміну")
        else:
            bot.answer_callback_query(call.id, "Ок")
        advance(); return

    if call.data.startswith("trp_"):
        idx = int(call.data[4:])
        cands = st.get('cands', [])
        if idx >= len(cands):
            bot.answer_callback_query(call.id, "Застаріло"); return
        new_name = cands[idx]
        save_orig = st.pop('ocr_new_original', None) or original
        if admin:
            if old_name:
                cache_ban_pair(original, old_name, cat)
                if cslug:
                    clients.client_cache_set_status(cslug, original, old_name, 'banned')
            cache_save(save_orig, {}, r.get('normalized', save_orig), new_name, cat, 100)
            cache_set_status(save_orig, new_name, 'confirmed')
            if save_orig != original:
                # Поки OCR читає по-старому — кеш ловить і старе прочитання
                cache_save(original, {}, save_orig, new_name, cat, 100)
                cache_set_status(original, new_name, 'confirmed')
            if cslug:
                clients.client_cache_save(cslug, save_orig, new_name, cat, 100)
                clients.client_cache_set_status(cslug, save_orig, new_name, 'confirmed')
            r['назва'] = new_name
            try:
                bot.edit_message_text(
                    f"✅ Навчено!\n{original[:40]}\n❌ {old_name[:50] or '—'}\n✅ {new_name[:60]}",
                    chat_id, st.get('msg_id', call.message.message_id))
            except Exception: pass
            bot.answer_callback_query(call.id, "✅ Збережено")
            suggest_knowledge_rule(chat_id, original, old_name, new_name)
        else:
            n = add_pending_fix({'original': original, 'old_name': old_name or None,
                                 'new_name': new_name, 'category': cat,
                                 'client_slug': cslug,
                                 'normalized': r.get('normalized',''),
                                 'user_id': call.from_user.id, 'username': uname})
            notify_admin_fix(uname, original, old_name, new_name, n)
            try:
                bot.edit_message_text(
                    f"📥 Надіслано адміну на підтвердження (черга: {n})\n"
                    f"{original[:40]}\n❌ {old_name[:50] or '—'}\n✅ {new_name[:60]}",
                    chat_id, st.get('msg_id', call.message.message_id))
            except Exception: pass
            bot.answer_callback_query(call.id, "📥 На розгляді")
        advance(); return

    # ═══ ЕТАП 1: 📖 Бот неправильно ПРОЧИТАВ (OCR-помилка) ═══
    if call.data == "tro":
        _manual_wait[chat_id] = {'mode': 'ocr_fix'}
        bot.answer_callback_query(call.id)
        bot.send_message(chat_id,
            f"📖 Бот прочитав з фото:\n«{original[:60]}»\n\n"
            f"✍️ Напиши як НАСПРАВДІ написано в списку\n"
            f"(бот запам'ятає почерк і одразу перешукає):",
            )
        return

    # ═══ ЕТАП 2: 🎯 Товар не той — кандидати з тих, що бот РОЗГЛЯДАВ ═══
    if call.data in ("trg", "trw", "trb", "trc"):
        if old_name and admin:
            cache_ban_pair(original, old_name, cat)
            if cslug:
                clients.client_cache_set_status(cslug, original, old_name, 'banned')
        seen = [c for c in (r.get('candidates_debug') or []) if c and c != old_name][:6]
        if not seen:
            # бот кандидатів не мав — одразу ручний ввід з бази
            _manual_wait[chat_id] = {'mode': 'train'}
            bot.answer_callback_query(call.id)
            bot.send_message(chat_id,
                "✍️ Бот не мав кандидатів. Напиши назву товару з бази (можна частину):")
            return
        st['cands'] = seen
        mk = InlineKeyboardMarkup(row_width=1)
        for i2, name in enumerate(seen):
            mk.add(InlineKeyboardButton(f"{i2+1}. {name[:55]}", callback_data=f"trp_{i2}"))
        mk.add(InlineKeyboardButton("✍️ Немає тут — ввести з бази", callback_data="trm"))
        mk.add(InlineKeyboardButton("❌ Немає правильного (бан)",     callback_data="trn"))
        hdr = "❌ Забанено" if admin else "❌ Позначено"
        try:
            bot.edit_message_text(
                f"🎯 Рядок {row}: {original[:45]}\n{hdr}: {old_name[:55] or '—'}\n\n"
                f"Бот розглядав ці варіанти — тапни ПРАВИЛЬНИЙ:",
                chat_id, st.get('msg_id', call.message.message_id), reply_markup=mk)
        except Exception:
            m2 = bot.send_message(chat_id, "Тапни правильний:", reply_markup=mk)
            st['msg_id'] = m2.message_id
        bot.answer_callback_query(call.id)
        return

    # (legacy-блок нижче не досяжний для нових кнопок)
    # Причина (trw/trb/trc): адмін банить одразу, користувач — при виборі нового
    if old_name and admin:
        cache_ban_pair(original, old_name, cat)
        if cslug:
            clients.client_cache_set_status(cslug, original, old_name, 'banned')
    query = r.get('normalized') or original
    cands = [c['name'] for c in keyword_search(query, top_n=10) if c['name'] != old_name]
    if call.data == "trb" and old_name:
        _ob = ''
        for _k, _t in BRAND_TOKENS.items():
            if any(x.lower() in old_name.lower() for x in _t):
                _ob = _t[0].lower(); break
        if _ob:
            _f = [c for c in cands if _ob not in c.lower()]
            cands = _f or cands
    cands = cands[:7]
    st['cands'] = cands
    mk = InlineKeyboardMarkup(row_width=1)
    for i2, name in enumerate(cands):
        mk.add(InlineKeyboardButton(f"{i2+1}. {name[:55]}", callback_data=f"trp_{i2}"))
    mk.add(InlineKeyboardButton("✍️ Ввести назву вручну", callback_data="trm"))
    mk.add(InlineKeyboardButton("❌ Немає правильного", callback_data="trn"))
    hdr = "❌ Забанено" if admin else "❌ Позначено як помилку"
    try:
        bot.edit_message_text(
            f"🎓 Рядок {row}: {original[:45]}\n{hdr}: {old_name[:55] or '—'}\n\nТапни ПРАВИЛЬНИЙ:",
            chat_id, st.get('msg_id', call.message.message_id), reply_markup=mk)
    except Exception:
        m2 = bot.send_message(chat_id, "Тапни правильний варіант:", reply_markup=mk)
        st['msg_id'] = m2.message_id
    bot.answer_callback_query(call.id)


@bot.message_handler(func=lambda m: m.text and re.match(r'^вірно\s+\d+', m.text.lower()))
def handle_virno(message):
    row = int(re.search(r'\d+', message.text).group())
    last = last_results.get(message.chat.id)
    if not last:
        bot.reply_to(message, "⚠️ Немає замовлення в пам'яті."); return
    if row < 1 or row > len(last['результати']):
        bot.reply_to(message, f"⚠️ Рядок {row} не існує."); return
    r = last['результати'][row-1]
    original, назва, cat = r.get('original',''), r.get('назва',''), r.get('category','other')
    if not назва:
        bot.reply_to(message, "⚠️ Цей рядок не знайдено — використай `виправ N`", parse_mode="Markdown"); return
    if is_admin(message.from_user.id):
        if not cache_set_status(original, назва, 'confirmed'):
            cache_save(original, {}, r.get('normalized', original), назва, cat, 100)
            cache_set_status(original, назва, 'confirmed')
        if last.get('client_slug'):
            clients.client_cache_save(last['client_slug'], original, назва, cat, 100)
            clients.client_cache_set_status(last['client_slug'], original, назва, 'confirmed')
        bot.reply_to(message, f"✅ Рядок {row} підтверджено (завжди так):\n{назва[:60]}")
    else:
        uname = message.from_user.username or str(message.from_user.id)
        n = add_pending_fix({'original': original, 'old_name': None, 'new_name': назва,
                             'category': cat, 'client_slug': last.get('client_slug'),
                             'normalized': r.get('normalized',''),
                             'user_id': message.from_user.id, 'username': uname})
        notify_admin_fix(uname, original, None, назва, n)
        bot.reply_to(message, f"📥 Підтвердження рядка {row} надіслано адміну (черга: {n})")


@bot.message_handler(func=lambda m: m.text and re.match(r'^помилка\s+\d+', m.text.lower()))
def handle_pomylka(message):
    row = int(re.search(r'\d+', message.text).group())
    last = last_results.get(message.chat.id)
    if not last:
        bot.reply_to(message, "⚠️ Немає замовлення в пам'яті."); return
    if row < 1 or row > len(last['результати']):
        bot.reply_to(message, f"⚠️ Рядок {row} не існує."); return
    r = last['результати'][row-1]
    original, назва, cat = r.get('original',''), r.get('назва',''), r.get('category','other')
    if not назва:
        bot.reply_to(message, "⚠️ Рядок і так не знайдено."); return
    if is_admin(message.from_user.id):
        cache_ban_pair(original, назва, cat)
        if last.get('client_slug'):
            clients.client_cache_set_status(last['client_slug'], original, назва, 'banned')
        bot.reply_to(message, f"❌ Рядок {row} забанено (ніколи так):\n{назва[:60]}")
    else:
        uname = message.from_user.username or str(message.from_user.id)
        n = add_pending_fix({'original': original, 'old_name': назва, 'new_name': None,
                             'category': cat, 'client_slug': last.get('client_slug'),
                             'normalized': r.get('normalized',''),
                             'user_id': message.from_user.id, 'username': uname})
        notify_admin_fix(uname, original, назва, None, n)
        bot.reply_to(message, f"📥 Позначку помилки рядка {row} надіслано адміну (черга: {n})")


# ═══════════════════════════════════════════════════════════════════════════════
# КОМАНДИ КЛІЄНТІВ
# ═══════════════════════════════════════════════════════════════════════════════
@bot.message_handler(func=lambda m: m.text and m.text.lower().startswith('новий клієнт'))
def handle_new_client(message):
    rest = message.text[12:].strip()
    if not rest:
        bot.reply_to(message, "Формат: `новий клієнт Петренко, примітка`", parse_mode="Markdown")
        return
    parts = rest.split(',', 1)
    name = parts[0].strip()
    notes = parts[1].strip() if len(parts) > 1 else ""
    ok, result = clients.create_client(name, notes)
    if ok:
        clients.set_active(message.chat.id, result)
        bot.reply_to(message, f"✅ Створено *{name}* і активовано. Кидай фото!", parse_mode="Markdown")
    else:
        bot.reply_to(message, f"⚠️ {result}")


@bot.message_handler(func=lambda m: m.text and m.text.lower().strip() == 'клієнти')
def handle_clients_list(message):
    index = clients.list_clients()
    if not index:
        bot.reply_to(message, "📁 Клієнтів немає.\n`новий клієнт <ім'я>`", parse_mode="Markdown")
        return
    lines = []
    for slug, name in sorted(index.items(), key=lambda x: x[1]):
        p = clients.get_profile(slug)
        lines.append(f"• {name} ({p.get('orders_count',0) if p else 0} зам.)")
    bot.reply_to(message, f"📁 Клієнти ({len(index)}):\n" + "\n".join(lines))


@bot.message_handler(func=lambda m: m.text and m.text.lower().startswith('клієнт'))
def handle_client(message):
    rest = message.text[6:].strip()
    if not rest:
        slug = clients.get_active(message.chat.id)
        if slug:
            p = clients.get_profile(slug)
            bot.reply_to(message, f"👤 Активний: *{p['name'] if p else slug}*\n`клієнт стоп` — скинути", parse_mode="Markdown")
        else:
            bot.reply_to(message, "Немає активного.\n`клієнт <ім'я>` — активувати", parse_mode="Markdown")
        return
    if rest.lower() in ('стоп', 'скинути', 'off'):
        clients.clear_active(message.chat.id)
        bot.reply_to(message, "✅ Клієнта скинуто.")
        return
    if ':' in rest:
        name_part, note = rest.split(':', 1)
        slug = clients.find_client(name_part.strip())
        if not slug:
            bot.reply_to(message, f"⚠️ '{name_part.strip()}' не знайдено.")
            return
        clients.add_note(slug, note.strip())
        bot.reply_to(message, "✅ Примітку додано.")
        return
    slug = clients.find_client(rest)
    if not slug:
        bot.reply_to(message, f"⚠️ '{rest}' не знайдено.\n`новий клієнт {rest}`", parse_mode="Markdown")
        return
    clients.set_active(message.chat.id, slug)
    p = clients.get_profile(slug)
    prefs = clients.get_preferences(slug)
    top = ", ".join(b for b, _ in prefs.get('top_brands', [])[:3]) or "ще немає даних"
    notes = p.get('notes', []) if p else []
    notes_s = "\n".join(f"  • {n}" for n in notes[-3:] if n) or "  —"
    bot.reply_to(message,
        f"✅ Активовано: *{p['name'] if p else slug}*\n"
        f"📦 Замовлень: {p.get('orders_count',0) if p else 0}\n"
        f"🏷 Топ виробники: {top}\n📝 Примітки:\n{notes_s}\n\nКидай фото!",
        parse_mode="Markdown")


# ═══════════════════════════════════════════════════════════════════════════════
# ВИПРАВ N [= текст] — навчання одним тапом
# ═══════════════════════════════════════════════════════════════════════════════
@bot.message_handler(func=lambda m: m.text and re.match(r'^виправ\s+\d+', m.text.lower().strip()))
def handle_fix(message):
    m = re.match(r'^виправ\s+(\d+)(?:\s*=\s*(.+))?$', message.text.strip(), re.IGNORECASE)
    row = int(m.group(1))
    manual = (m.group(2) or '').strip()
    last = last_results.get(message.chat.id)
    if not last or not last['результати']:
        bot.reply_to(message, "⚠️ Немає замовлення в пам'яті.")
        return
    if row < 1 or row > len(last['результати']):
        bot.reply_to(message, f"⚠️ Рядок {row} не існує (всього {len(last['результати'])}).")
        return
    r = last['результати'][row-1]
    query = manual or r.get('normalized') or r.get('original','')
    cur = r.get('назва','')
    cands = [c['name'] for c in keyword_search(query, top_n=9) if c['name'] != cur][:8]
    if not cands:
        bot.reply_to(message, "😕 Кандидатів немає. Спробуй: `виправ N = інший текст`", parse_mode="Markdown")
        return
    _fix_state[message.chat.id] = {'row': row, 'cands': cands}
    mk = InlineKeyboardMarkup(row_width=1)
    for i, name in enumerate(cands):
        mk.add(InlineKeyboardButton(f"{i+1}. {name[:55]}", callback_data=f"fx_{i}"))
    mk.add(InlineKeyboardButton("✍️ Ввести назву вручну", callback_data="fx_man"))
    mk.add(InlineKeyboardButton("❌ Немає правильного (тільки бан)", callback_data="fx_ban"))
    bot.reply_to(message,
        f"🎓 Рядок {row}: `{r.get('original','')[:45]}`\n"
        f"Зараз: {cur[:60] or '(не знайдено)'}\n\nТапни ПРАВИЛЬНИЙ:",
        parse_mode="Markdown", reply_markup=mk)


@bot.callback_query_handler(func=lambda c: c.data.startswith('fx_'))
def handle_fix_pick(call):
    if call.data == "fx_man":
        st0 = _fix_state.get(call.message.chat.id)
        if not st0:
            bot.answer_callback_query(call.id, "Сесія застаріла"); return
        _manual_wait[call.message.chat.id] = {'mode': 'fix'}
        bot.answer_callback_query(call.id)
        bot.send_message(call.message.chat.id,
            "✍️ Напиши назву товару (як у прайсі, можна частину):\n"
            "напр: `коліно 110 45 остендорф`", parse_mode="Markdown")
        return
    st = _fix_state.pop(call.message.chat.id, None)
    last = last_results.get(call.message.chat.id)
    if not st or not last:
        bot.answer_callback_query(call.id, "Сесія застаріла")
        return
    admin = is_admin(call.from_user.id)
    r = last['результати'][st['row']-1]
    original = r.get('original','')
    old_name = r.get('назва','')
    cat = r.get('category','other')
    cslug = last.get('client_slug')
    uname = call.from_user.username or str(call.from_user.id)

    if call.data == "fx_ban":
        if not old_name:
            bot.answer_callback_query(call.id, "Нема чого банити"); return
        if admin:
            cache_ban_pair(original, old_name, cat)
            if cslug:
                clients.client_cache_set_status(cslug, original, old_name, 'banned')
            bot.edit_message_text(f"❌ Забанено: {original[:40]} → {old_name[:50]}",
                call.message.chat.id, call.message.message_id)
            bot.answer_callback_query(call.id, "Забанено")
        else:
            n = add_pending_fix({'original': original, 'old_name': old_name, 'new_name': None,
                                 'category': cat, 'client_slug': cslug,
                                 'normalized': r.get('normalized',''),
                                 'user_id': call.from_user.id, 'username': uname})
            notify_admin_fix(uname, original, old_name, None, n)
            bot.edit_message_text(f"📥 Надіслано адміну (черга: {n}):\nзаборонити {old_name[:50]}",
                call.message.chat.id, call.message.message_id)
            bot.answer_callback_query(call.id, "📥 На розгляді")
        return

    idx = int(call.data[3:])
    if idx >= len(st.get('cands', [])):
        bot.answer_callback_query(call.id, "Застаріло"); return
    new_name = st['cands'][idx]
    if admin:
        if old_name:
            cache_ban_pair(original, old_name, cat)
            if cslug:
                clients.client_cache_set_status(cslug, original, old_name, 'banned')
        cache_save(original, {}, r.get('normalized', original), new_name, cat, 100)
        cache_set_status(original, new_name, 'confirmed')
        if cslug:
            clients.client_cache_save(cslug, original, new_name, cat, 100)
            clients.client_cache_set_status(cslug, original, new_name, 'confirmed')
        r['назва'] = new_name
        bot.edit_message_text(
            f"✅ Навчено!\n{original[:40]}\n❌ {old_name[:50] or '—'}\n✅ {new_name[:60]}",
            call.message.chat.id, call.message.message_id)
        bot.answer_callback_query(call.id, "Збережено")
        # 💡 Бот сам формулює правило з цього виправлення
        suggest_knowledge_rule(call.message.chat.id, original, old_name, new_name)
        # Пропонуємо виправити OCR якщо original і new_name сильно різняться
        orig_words = set(re.findall(r'[а-яёіїєґa-z]+', original.lower()))
        new_words = set(re.findall(r'[а-яёіїєґa-z]+', new_name.lower()))
        ocr_diff = orig_words - new_words - {'шт','м','мп','компл','рул','пог'}
        if ocr_diff and admin:
            bot.send_message(call.message.chat.id,
                f"🔤 OCR-помилка? Якщо Gemini неправильно прочитав слово — виправ:\n"
                f"Слова з запиту яких немає в результаті: `{'`, `'.join(list(ocr_diff)[:4])}`\n\n"
                f"Виправити: `ocr <неправильне> = <правильне>`\nПриклад: `ocr кільця = гільза`",
                parse_mode="Markdown")
    else:
        n = add_pending_fix({'original': original, 'old_name': old_name or None,
                             'new_name': new_name, 'category': cat, 'client_slug': cslug,
                             'normalized': r.get('normalized',''),
                             'user_id': call.from_user.id, 'username': uname})
        notify_admin_fix(uname, original, old_name, new_name, n)
        bot.edit_message_text(
            f"📥 Надіслано адміну на підтвердження (черга: {n})\n"
            f"{original[:40]}\n❌ {old_name[:50] or '—'}\n✅ {new_name[:60]}",
            call.message.chat.id, call.message.message_id)
        bot.answer_callback_query(call.id, "📥 На розгляді")


@bot.message_handler(func=lambda m: m.text and m.text.lower().startswith('ocr '))
def handle_ocr_correction(message):
    """
    Корекція OCR: ocr кільця = гільза
    Після цього Gemini при читанні "кільця" завжди виправить на "гільза".
    """
    rest = message.text[4:].strip()
    if '=' not in rest:
        bot.reply_to(message,
            "Формат: `ocr <неправильно> = <правильно>`\n"
            "Приклад: `ocr кільця = гільза`\n"
            "Приклад: `ocr 2S = 25`\n"
            "Дивитись збережені: `ocr список`",
            parse_mode="Markdown")
        return
    parts = rest.split('=', 1)
    wrong = parts[0].strip()
    right = parts[1].strip()
    if not wrong or not right:
        bot.reply_to(message, "⚠️ Вкажи і неправильне і правильне слово.")
        return
    save_ocr_correction(wrong, right)
    bot.reply_to(message,
        f"✅ OCR корекцію збережено:\n`{wrong}` → `{right}`\n\n"
        f"Тепер Gemini виправлятиме при читанні фото і PDF.",
        parse_mode="Markdown")


@bot.message_handler(func=lambda m: m.text and m.text.lower().strip() == 'ocr список')
def handle_ocr_list(message):
    d = load_ocr_corrections()
    if not d:
        bot.reply_to(message, "📋 Корекцій OCR поки немає.\nДодати: `ocr кільця = гільза`",
                     parse_mode="Markdown")
        return
    lines = [f"  `{w}` → `{r}`" for w, r in d.items()]
    bot.reply_to(message, f"📋 OCR корекції ({len(d)}):\n" + "\n".join(lines),
                 parse_mode="Markdown")


@bot.message_handler(func=lambda m: m.text and m.text.lower().startswith('правило'))
def handle_rule(message):
    rule = message.text[7:].strip()
    if not rule:
        bot.reply_to(message, "Напиши правило після слова 'правило'."); return
    if is_admin(message.from_user.id):
        add_rule(rule)
        bot.reply_to(message, f"✅ Записав:\n_{rule}_", parse_mode="Markdown")
    else:
        rules = load_pending_rules()
        rules.append({"rule": rule, "user_id": message.from_user.id,
                      "username": message.from_user.username or str(message.from_user.id),
                      "date": time.strftime("%Y-%m-%d %H:%M")})
        save_pending_rules(rules)
        bot.reply_to(message, f"📥 Правило відправлено на розгляд ({len(rules)} в черзі):\n_{rule}_",
                     parse_mode="Markdown")
        try:
            bot.send_message(ADMIN_ID, f"🔔 Нове правило від @{message.from_user.username}:\n`{rule}`",
                             parse_mode="Markdown")
        except Exception: pass


@bot.message_handler(commands=['кеш', 'cache'])
def handle_cache_info(message):
    cache = get_cache()
    if not cache:
        bot.reply_to(message, "📋 Кеш порожній — заповниться після замовлень.")
        return
    icons = {'confirmed':'✅','banned':'❌','auto':'🔹'}
    lines = []
    for k, v in list(cache.items())[-8:]:
        orig = k.split("::")[0][:35]
        lines.append(f"{icons.get(v.get('status','auto'),'🔹')} `{orig}` → {v.get('catalog_name','')[:45]}")
    bot.reply_to(message,
        f"📋 Кеш: *{len(cache)}* записів\n✅ підтв | ❌ бан | 🔹 авто\n\n" + "\n".join(lines),
        parse_mode="Markdown")


@bot.message_handler(content_types=['photo'])
def handle_photo(message):
    fuid = message.photo[-1].file_unique_id
    _b = user_batches.get(message.chat.id)
    if _b and any(it.get('fuid') == fuid for it in _b.get('items', [])):
        return
    for attempt in range(3):
        try:
            file_info = bot.get_file(message.photo[-1].file_id)
            downloaded = bot.download_file(file_info.file_path)
            image_b64 = base64.b64encode(downloaded).decode('utf-8')
            caption = message.caption or ""
            hint = pending_hints.pop(message.chat.id, "")
            full_caption = " | ".join(filter(None, [caption, hint]))
            add_to_batch(message.chat.id, {
                'type': 'photo', 'data': image_b64, 'caption': full_caption,
                'fuid': fuid,
                'username': message.from_user.username or str(message.from_user.id),
            })
            return
        except Exception as e:
            if attempt == 2:
                bot.reply_to(message, f"❌ Не вдалося завантажити фото: {e}")
            else:
                time.sleep(2)


@bot.message_handler(content_types=['document'])
def handle_document(message):
    doc = message.document
    mime = (doc.mime_type or '').lower()
    fname = (doc.file_name or '').lower()
    is_pdf = mime == 'application/pdf' or fname.endswith('.pdf')
    if not (is_pdf or mime.startswith('image/') or fname.endswith(('.jpg','.jpeg','.png','.webp'))):
        bot.reply_to(message, "📎 Приймаю фото (jpg/png) або PDF-специфікацію.")
        return
    fuid = doc.file_unique_id
    _b = user_batches.get(message.chat.id)
    if _b and any(it.get('fuid') == fuid for it in _b.get('items', [])):
        return
    for attempt in range(3):
        try:
            file_info = bot.get_file(doc.file_id)
            downloaded = bot.download_file(file_info.file_path)
            image_b64 = base64.b64encode(downloaded).decode('utf-8')
            caption = message.caption or ""
            hint = pending_hints.pop(message.chat.id, "")
            full_caption = " | ".join(filter(None, [caption, hint]))
            add_to_batch(message.chat.id, {
                'type': 'pdf' if is_pdf else 'photo',
                'data': image_b64, 'caption': full_caption, 'fuid': fuid,
                'username': message.from_user.username or str(message.from_user.id),
            })
            return
        except Exception as e:
            if attempt == 2:
                bot.reply_to(message, f"❌ Не вдалося завантажити: {e}")
            else:
                time.sleep(2)


@bot.message_handler(func=lambda m: m.text and m.text.lower().startswith('пошук'))
def handle_text_search(message):
    запит = message.text[5:].strip()
    if запит:
        hint = pending_hints.pop(message.chat.id, "")
        add_to_batch(message.chat.id, {
            'type': 'text', 'text': запит, 'caption': hint,
            'username': message.from_user.username or str(message.from_user.id),
        })
    else:
        bot.reply_to(message, "Напиши запит після слова 'пошук'.")


@bot.message_handler(commands=['stop'])
def handle_stop(message):
    chat_id = message.chat.id
    stop_flags[chat_id] = True
    if chat_id in user_batches:
        if 'timer' in user_batches[chat_id]:
            user_batches[chat_id]['timer'].cancel()
        user_batches.pop(chat_id, None)
    bot.reply_to(message, "🛑 Зупинено.")


@bot.message_handler(func=lambda m: m.text and m.chat.id in _manual_wait)
def handle_manual_input(message):
    """Отримуємо ручну назву товару для виправ/навчання."""
    state = _manual_wait.pop(message.chat.id, None)
    if not state:
        return
    mode = state.get('mode')
    query = message.text.strip()
    cands = [c['name'] for c in keyword_search(query, top_n=9)]
    if not cands:
        bot.reply_to(message, "😕 Нічого не знайдено. Спробуй іншу назву або частину.")
        return
    if mode == 'fix':
        st = _fix_state.get(message.chat.id)
        if not st:
            bot.reply_to(message, "⚠️ Сесія виправлення завершена.")
            return
        st['cands'] = cands
        mk = InlineKeyboardMarkup(row_width=1)
        for i, name in enumerate(cands):
            mk.add(InlineKeyboardButton(f"{i+1}. {name[:55]}", callback_data=f"fx_{i}"))
        mk.add(InlineKeyboardButton("✍️ Ввести іншу назву", callback_data="fx_man"))
        mk.add(InlineKeyboardButton("❌ Немає правильного (тільки бан)", callback_data="fx_ban"))
        bot.reply_to(message, f"🔍 Знайдено за '{query}':\nТапни правильний:", reply_markup=mk)
    elif mode == 'ocr_fix':
        # ЕТАП 1: користувач написав як НАСПРАВДІ було в списку
        st = _train_state.get(message.chat.id)
        last = last_results.get(message.chat.id)
        if not st or not last:
            bot.reply_to(message, "⚠️ Сесія навчання завершена."); return
        row = st['rows'][st['i']]
        r = last['результати'][row-1]
        old_original = r.get('original','')
        corrected = query

        # 1) Авто-виявлення OCR-пар: слова що відрізняються
        w_re = r'[а-яёіїєґa-z]+|\d+(?:[.,/]\d+)?'
        old_w = re.findall(w_re, old_original.lower())
        new_w = re.findall(w_re, corrected.lower())
        wrongs = [w for w in old_w if w not in new_w and len(w) > 1]
        rights = [w for w in new_w if w not in old_w and len(w) > 1]
        pairs = list(zip(wrongs, rights))[:3]
        if pairs:
            st['ocr_pairs'] = pairs
            mk = InlineKeyboardMarkup(row_width=1)
            for pi, (w, rt) in enumerate(pairs):
                mk.add(InlineKeyboardButton(f"🔤 Запам'ятати: «{w}» → «{rt}»",
                                            callback_data=f"ocrs_{pi}"))
            bot.send_message(message.chat.id,
                "Схоже на помилки почерку — збережу корекції для наступних фото?",
                reply_markup=mk)

        # 2) Одразу перешукуємо по ПРАВИЛЬНОМУ тексту
        псевдо = {'original': corrected, 'normalized': corrected,
                  'category': r.get('category','other')}
        cands = [c['name'] for c in smart_search(псевдо, top_n=8)]
        cands = [c for c in cands if c != r.get('назва','')][:7]
        if not cands:
            bot.reply_to(message, "😕 По новому тексту нічого. Спробуй ще раз або точнішу назву.")
            _manual_wait[message.chat.id] = {'mode': 'ocr_fix'}
            return
        st['cands'] = cands
        st['ocr_new_original'] = corrected   # confirm піде на ПРАВИЛЬНИЙ ключ
        mk = InlineKeyboardMarkup(row_width=1)
        for i, name in enumerate(cands):
            mk.add(InlineKeyboardButton(f"{i+1}. {name[:55]}", callback_data=f"trp_{i}"))
        mk.add(InlineKeyboardButton("✍️ Ввести з бази точніше", callback_data="trm"))
        mk.add(InlineKeyboardButton("❌ Немає правильного",      callback_data="trn"))
        bot.reply_to(message,
            f"🔍 Пошук за «{corrected[:45]}»:\nТапни ПРАВИЛЬНИЙ товар:", reply_markup=mk)
    elif mode == 'train':
        st = _train_state.get(message.chat.id)
        if not st:
            bot.reply_to(message, "⚠️ Сесія навчання завершена.")
            return
        st['cands'] = cands
        mk = InlineKeyboardMarkup(row_width=1)
        for i, name in enumerate(cands):
            mk.add(InlineKeyboardButton(f"{i+1}. {name[:55]}", callback_data=f"trp_{i}"))
        mk.add(InlineKeyboardButton("✍️ Ввести іншу назву", callback_data="trm"))
        mk.add(InlineKeyboardButton("❌ Немає правильного", callback_data="trn"))
        bot.reply_to(message, f"🔍 Знайдено за '{query}':\nТапни правильний:", reply_markup=mk)


@bot.message_handler(func=lambda m: m.text and not m.text.startswith('/')
                     and not m.text.lower().startswith('пошук')
                     and not m.text.lower().startswith('правило')
                     and not re.match(r'^(вірно|помилка|виправ)\s+\d+', m.text.lower().strip())
                     and not m.text.lower().startswith('клієнт')
                     and not m.text.lower().startswith('новий клієнт')
                     and not m.text.startswith(('📸','🛑','📋','📊','👥','👑')))
def handle_text_hint(message):
    text = message.text.strip()
    if not text:
        return

    # Детект текст-списку: ≥3 рядки АБО ≥2 рядки з "число+одиниця"
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    has_qty = sum(1 for l in lines if re.search(r'\d+\s*(шт|м\b|пог|компл|рул)', l.lower()))
    is_list = len(lines) >= 3 or has_qty >= 2

    if is_list:
        # Пропонуємо вибір: підібрати як список або зберегти як підказку
        _text_pending[message.chat.id] = text
        mk = InlineKeyboardMarkup()
        mk.add(InlineKeyboardButton("🔍 Підібрати список", callback_data="txts"),
               InlineKeyboardButton("💬 Це підказка до фото", callback_data="txth"))
        bot.reply_to(message,
            f"Схоже на список ({len(lines)} рядків). Що робити?",
            reply_markup=mk)
    else:
        # Коротка підказка — зберігаємо до наступного фото
        pending_hints[message.chat.id] = text
        bot.reply_to(message, f"💬 Підказка: _{text}_\nТепер кидай фото!", parse_mode="Markdown")
        def clear(cid): pending_hints.pop(cid, None)
        t = threading.Timer(120.0, clear, args=[message.chat.id])
        t.daemon = True
        t.start()


@bot.callback_query_handler(func=lambda c: c.data in ("txts", "txth"))
def handle_text_choice(call):
    chat_id = call.message.chat.id
    text = _text_pending.pop(chat_id, None)
    if not text:
        bot.answer_callback_query(call.id, "Сесія застаріла")
        return
    if call.data == "txts":
        # Підібрати як список
        add_to_batch(chat_id, {
            'type': 'text', 'text': text, 'caption': '',
            'username': call.from_user.username or str(call.from_user.id),
        })
        try:
            bot.edit_message_text("✅ Прийнято! Обробляю список...",
                chat_id, call.message.message_id)
        except Exception:
            pass
    else:
        # Зберегти як підказку до наступного фото
        pending_hints[chat_id] = text
        try:
            bot.edit_message_text(f"💬 Підказка збережена.\nТепер кидай фото!",
                chat_id, call.message.message_id)
        except Exception:
            pass
    bot.answer_callback_query(call.id)


try:
    storage.start_autosave(60)
except Exception as _e:
    print(f"⚠️ storage autosave: {_e}", flush=True)

print("🤖 bot.py завантажено, хендлери зареєстровані", flush=True)

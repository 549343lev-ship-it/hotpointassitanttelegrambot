"""
synonym_export.py — Генерує Excel-словник синонімів.

Джерела:
  - normalization_cache.json (глобальний кеш)
  - clients/{slug}/cache.json (опційно, якщо include_clients=True)

Колонки:
  A: Оригінал     B: Нормалізована назва    C: Товар у каталозі
  D: Категорія    E: Джерело                F: Статус
  G: Впевненість  H: Дата                   I: Клієнт (якщо є)
"""

import os
import re
import json
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config.settings import DATA_DIR

CACHE_FILE  = os.path.join(DATA_DIR, "normalization_cache.json")
CLIENTS_DIR = os.path.join(DATA_DIR, "clients")

HEADERS = [
    "Оригінал (рукопис)",
    "Нормалізована назва",
    "Товар у каталозі",
    "Категорія",
    "Джерело",
    "Статус",
    "Впевненість, %",
    "Дата",
    "Клієнт",
]
COL_WIDTHS = [35, 38, 45, 22, 14, 13, 14, 13, 18]

# ── Джерела ───────────────────────────────────────────────────────────────────
SOURCE_LABEL = {
    "auto":      "🤖 Авто",
    "training":  "🎓 Навчання",
    "confirmed": "✅ Підтверджено",
    "client":    "👤 Клієнт",
}

STATUS_ICON = {
    "confirmed": "✅",
    "auto":      "🤖",
    "banned":    "🚫",
}

# ── Кольори рядків ────────────────────────────────────────────────────────────
ROW_FILLS = {
    ("training",  "confirmed"): PatternFill("solid", fgColor="C6EFCE"),  # яскраво-зелений
    ("confirmed", "confirmed"): PatternFill("solid", fgColor="D6F5D6"),  # світло-зелений
    ("auto",      "auto"):      PatternFill("solid", fgColor="EBF3FB"),  # блакитний
    ("client",    "confirmed"): PatternFill("solid", fgColor="FFF2CC"),  # жовтий
    ("client",    "auto"):      PatternFill("solid", fgColor="FFF9E6"),  # світло-жовтий
    ("auto",      "banned"):    PatternFill("solid", fgColor="FDDEDE"),  # червоний
    ("confirmed", "banned"):    PatternFill("solid", fgColor="FDDEDE"),
    ("training",  "banned"):    PatternFill("solid", fgColor="FDDEDE"),
}
DEFAULT_FILL    = PatternFill("solid", fgColor="F5F5F5")
HEADER_FILL     = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT     = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CELL_FONT       = Font(name="Arial", size=9)
TRAINING_FONT   = Font(name="Arial", size=9, bold=True)  # навчання виділяємо жирним
THIN_BORDER     = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


# ── Утиліти ───────────────────────────────────────────────────────────────────

def _load_json(path: str) -> dict:
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _parse_original(raw_key: str) -> str:
    """'оригінал::brands::ban1' → 'оригінал'"""
    key = raw_key.split("::", 1)[0]
    key = re.sub(r'::ban\d+$', '', key)
    return key.strip()


def _is_ban_marker(raw_key: str) -> bool:
    """Пропускаємо суфіксні бан-записи ::ban1, ::ban2 ..."""
    parts = raw_key.split("::")
    return len(parts) > 1 and re.match(r'^ban\d+$', parts[-1]) is not None


def _collect_global(raw: dict) -> list[dict]:
    rows = []
    for raw_key, entry in raw.items():
        if _is_ban_marker(raw_key):
            continue
        source = entry.get("source", "auto")
        rows.append({
            "original":     _parse_original(raw_key),
            "normalized":   entry.get("normalized", ""),
            "catalog_name": entry.get("catalog_name", ""),
            "category":     entry.get("category", ""),
            "source":       source,
            "status":       entry.get("status", "auto"),
            "confidence":   entry.get("confidence", ""),
            "saved_at":     entry.get("saved_at", ""),
            "client":       "",
        })
    return rows


def _collect_clients() -> list[dict]:
    """Збирає унікальні confirmed-записи з усіх клієнтських кешів."""
    rows = []
    if not os.path.isdir(CLIENTS_DIR):
        return rows
    seen: set[tuple] = set()  # (original, catalog_name) — дедуплікація
    for slug in os.listdir(CLIENTS_DIR):
        cache_path = os.path.join(CLIENTS_DIR, slug, "cache.json")
        profile_path = os.path.join(CLIENTS_DIR, slug, "profile.json")
        raw = _load_json(cache_path)
        profile = _load_json(profile_path)
        client_name = profile.get("name", slug)
        for raw_key, entry in raw.items():
            if _is_ban_marker(raw_key):
                continue
            status = entry.get("status", "auto")
            original     = re.sub(r'\s+', ' ', raw_key.split("::")[0]).strip()
            catalog_name = entry.get("catalog_name", "")
            dedup_key    = (original, catalog_name)
            if dedup_key in seen:
                continue
            seen.add(dedup_key)
            rows.append({
                "original":     original,
                "normalized":   entry.get("catalog_name", ""),  # у клієнт-кеші normalized немає
                "catalog_name": catalog_name,
                "category":     entry.get("category", ""),
                "source":       "client",
                "status":       status,
                "confidence":   entry.get("confidence", ""),
                "saved_at":     entry.get("saved_at", ""),
                "client":       client_name,
            })
    return rows


def _sort_key(row: dict) -> tuple:
    source_order = {"training": 0, "confirmed": 1, "auto": 2, "client": 3}
    status_order = {"confirmed": 0, "auto": 1, "banned": 2}
    return (
        source_order.get(row["source"], 9),
        status_order.get(row["status"], 9),
        row.get("category", ""),
    )


# ── Основна функція ───────────────────────────────────────────────────────────

def build_synonym_excel(output_path: str, include_clients: bool = True) -> int:
    """
    Будує Excel-словник синонімів.
    include_clients=True — додає окремий аркуш з клієнтськими кешами.
    Повертає кількість рядків глобального кешу.
    """
    global_raw    = _load_json(CACHE_FILE)
    global_rows   = _collect_global(global_raw)
    client_rows   = _collect_clients() if include_clients else []

    wb = Workbook()

    # ── Аркуш 1: Глобальний словник ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Словник синонімів"
    _write_sheet(ws, sorted(global_rows, key=_sort_key), show_client_col=False)

    # ── Аркуш 2: Клієнтські записи (опційно) ─────────────────────────────────
    if include_clients and client_rows:
        ws2 = wb.create_sheet("Клієнтський кеш")
        _write_sheet(ws2, sorted(client_rows, key=_sort_key), show_client_col=True)

    # ── Аркуш 3: Статистика ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Статистика")
    _write_stats(ws3, global_rows, client_rows)

    wb.save(output_path)
    return len(global_rows)


def _write_sheet(ws, rows: list[dict], show_client_col: bool):
    headers = HEADERS if show_client_col else HEADERS[:-1]
    widths  = COL_WIDTHS if show_client_col else COL_WIDTHS[:-1]

    # Заголовок
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
    ws.row_dimensions[1].height = 22

    # Дані
    for row_num, row in enumerate(rows, 2):
        source = row["source"]
        status = row["status"]
        fill   = ROW_FILLS.get((source, status), DEFAULT_FILL)
        font   = TRAINING_FONT if source == "training" else CELL_FONT

        values = [
            row["original"],
            row["normalized"],
            row["catalog_name"],
            row["category"],
            SOURCE_LABEL.get(source, source),
            STATUS_ICON.get(status, status),
            row["confidence"],
            row["saved_at"],
        ]
        if show_client_col:
            values.append(row["client"])

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font      = font
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx <= 3))
        ws.row_dimensions[row_num].height = 18

    # Ширини + freeze + autofilter
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"


def _write_stats(ws, global_rows: list[dict], client_rows: list[dict]):
    ws.append(["Показник", "Значення"])
    ws["A1"].font = HEADER_FONT; ws["A1"].fill = HEADER_FILL
    ws["B1"].font = HEADER_FONT; ws["B1"].fill = HEADER_FILL

    def count(rows, **filters):
        return sum(1 for r in rows if all(r.get(k) == v for k, v in filters.items()))

    stats = [
        ("── ГЛОБАЛЬНИЙ КЕШ ──",              ""),
        ("Всього записів",                     len(global_rows)),
        ("🎓 З навчання",                      count(global_rows, source="training")),
        ("✅ Підтверджено (вірно N)",           count(global_rows, source="confirmed")),
        ("🤖 Авто (Voyage ≥95%)",              count(global_rows, source="auto")),
        ("🚫 Заблоковано",                     count(global_rows, status="banned")),
        ("── КЛІЄНТСЬКИЙ КЕШ ──",             ""),
        ("Унікальних записів (всі клієнти)",   len(client_rows)),
        ("  підтверджені",                     count(client_rows, status="confirmed")),
        ("  авто",                             count(client_rows, status="auto")),
        ("── ЗАГАЛЬНЕ ──",                     ""),
        ("Дата експорту",                      time.strftime("%Y-%m-%d %H:%M")),
    ]
    for label, value in stats:
        ws.append([label, value if value != "" else ""])

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14

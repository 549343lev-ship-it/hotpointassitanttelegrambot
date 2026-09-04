"""
excel_builder.py — Генерація Excel-файлу замовлення.

Вхід:  список результатів пошуку (з find_items)
Вихід: BytesIO з xlsx-файлом + списки незнайдених і сумнівних

Кольори:
  🟥 червоний (RED)   — не знайдено
  🟨 жовтий (YELLOW)  — знайдено, але сумнівно (низький confidence або fallback)
"""

import re
from io import BytesIO
import pandas as pd
from openpyxl.styles import PatternFill

from catalog.catalog import CATALOG

RED    = PatternFill('solid', fgColor='FFC7CE')   # заливка для незнайдених рядків
YELLOW = PatternFill('solid', fgColor='FFF3B0')   # заливка для сумнівних рядків


def parse_qty(s) -> tuple:   # розбиває рядок кількості на (число, одиниця): "10 шт" → (10, "шт")
    s = str(s or '').strip()
    if not s:
        return '', ''
    m = re.match(r'(\d+(?:[.,]\d+)?)\s*(.*)', s)
    if m:
        try:
            num = float(m.group(1).replace(',', '.'))
            num = int(num) if num == int(num) else num
        except Exception:
            num = m.group(1)
        return num, m.group(2).strip().rstrip('.').strip()
    return s, ''


def create_excel(результати: list[dict]) -> tuple[BytesIO, list, list]:  # будує xlsx-файл із результатів пошуку; повертає (файл, незнайдені, сумнівні)
    rows, flags = [], []
    not_found, warn = [], []

    for r in результати:
        if not r:
            continue
        conf     = r.get('confidence', 0)
        kw       = r.get('keyword_pct', 0)
        qty_num, qty_unit = parse_qty(r.get('qty', ''))

        # Товари що продаються по метражу
        import re as _re
        _norm = (r.get('normalized') or '').lower()
        _name = (r.get('назва', '') or '').lower()
        _orig = (r.get('original') or '').lower()

        # Плівка і демпферна — метраж з original ("плівка 50м")
        _is_film = any(x in _norm or x in _orig for x in ['плівк', 'фольг', 'розміт'])
        _is_damp = any(x in _norm or x in _orig for x in ['демпф', 'демпер'])
        if (_is_film or _is_damp) and qty_unit in ('шт', '', None):
            _lm2 = _re.search(r'(\d+(?:[.,]\d+)?)\s*м', _orig)
            if _lm2:
                _v2 = float(_lm2.group(1).replace(',', '.'))
                if _v2 >= 1:
                    qty_num  = int(_v2) if _v2 == int(_v2) else _v2
                    qty_unit = 'м'

        # PEX/PPR труба з L=Xм в назві: qty=1 рулон → qty=X метрів
        if qty_num == 1 and qty_unit in ('шт', '', None):
            _lm = _re.search(r'l=(\d+(?:[.,]\d+)?)\s*м', _norm)
            if not _lm:
                _lm = _re.search(r',\s*l\s*=\s*(\d+(?:[.,]\d+)?)\s*м', _name)
            if _lm:
                _v = float(_lm.group(1).replace(',', '.'))
                if _v > 1:
                    qty_num  = int(_v) if _v == int(_v) else _v
                    qty_unit = 'м'

        if r.get('знайдено'):
            suspicious = (
                conf < 70 or kw < 50
                or r.get('brand_warning')
                # підрядком, а не рівністю: джерело може мати суфікс
                # («⚠️ аналог 🌳» — спрацював фільтр гілки 1С)
                or any(s in r.get('джерело', '')
                       for s in ('fallback', 'вільний', 'аналог'))
            )
            # Маршрут = node_id від ROUTER (куди бот направляв пошук)
            _route_node = r.get('_node_id', '')   # ← від router через пос
            _pref       = r.get('_prefix', '')
            if _route_node and _route_node != 'xx':
                _маршрут = f"[{_route_node}]"
            else:
                _маршрут = f"[{_pref}]" if _pref else ''

            # Розділ = node_id ЗНАЙДЕНОГО товару (з каталогу)
            _found_item = r.get('_catalog_node', '')  # node_id товару що знайшло
            _розд = f"[{_found_item}]" if _found_item and _found_item != 'xx' else _маршрут

            # Бренд = який виробник використовувався при пошуку
            _brand = r.get('_used_brand', '') or r.get('brand_warning', '')

            rows.append({
                '№':            len(rows) + 1,
                'Артикул':      r.get('артикул', ''),
                'Наименование': r.get('назва_повна') or r.get('назва', ''),
                'Кількість':    qty_num,
                'Од.':          qty_unit,
                'Ціна':         r.get('ціна', ''),
                'Збіг':         f"🔍{kw}%/🤖{conf}%",
                'Джерело':      r.get('джерело', ''),
                'Маршрут':      _маршрут,
                'Розділ':       _розд,
                'Бренд':        _brand,
                'Чому знайшло': r.get('reason', ''),
                'Оригінал':     r.get('original', ''),
                'Нормалізовано': r.get('normalized', ''),
            })
            flags.append('warn' if suspicious else '')
            if suspicious:
                warn.append(r.get('original', ''))
        else:
            # для незнайдених — ставимо найближчий кандидат щоб менеджер міг оцінити
            cands = r.get('candidates_debug', [])
            best  = cands[0] if cands else ''
            art, full, price = '', best, ''
            for it in CATALOG:
                if it['name'] == best:
                    art   = it.get('artikul', '')
                    full  = it.get('name_full', best)
                    price = it.get('price', '')
                    break
            # Маршрут від router для незнайдених
            _node2 = r.get('_node_id', '')
            _pref2 = r.get('_prefix', '')
            if _node2 and _node2 != 'xx':
                _маршрут2 = f"[{_node2}]"
            else:
                _маршрут2 = f"[{_pref2}]" if _pref2 else ''

            rows.append({
                '№':            len(rows) + 1,
                'Артикул':      art,
                'Наименование': full,
                'Кількість':    qty_num,
                'Од.':          qty_unit,
                'Ціна':         price,
                'Збіг':         '—',
                'Джерело':      '❓ НЕ ЗНАЙДЕНО',
                'Маршрут':      _маршрут2,
                'Розділ':       '',
                'Бренд':        '',
                'Чому знайшло': (r.get('fail_reason', '') or '')[:100],
                'Оригінал':     r.get('original', ''),
                'Нормалізовано': r.get('normalized', ''),
            })
            flags.append('nf')
            not_found.append(r.get('original', ''))

    output = BytesIO()
    cols = ['№', 'Артикул', 'Наименование', 'Кількість', 'Од.',
            'Ціна', 'Збіг', 'Джерело', 'Маршрут', 'Розділ', 'Бренд',
            'Чому знайшло', 'Оригінал', 'Нормалізовано']

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=cols)
        df = df[cols]   # правильний порядок колонок
        df.to_excel(writer, index=False, sheet_name='Замовлення')
        ws = writer.sheets['Замовлення']
        # Ширини колонок: A=№, B=Артикул, C=Найменування, D=Кіл, E=Од,
        # F=Ціна, G=Збіг, H=Джерело, I=Маршрут, J=Розділ, K=Бренд,
        # L=Чому, M=Оригінал, N=Нормалізовано
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 55
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 16   # Маршрут
        ws.column_dimensions['J'].width = 16   # Розділ
        ws.column_dimensions['K'].width = 18   # Бренд
        ws.column_dimensions['L'].width = 20   # Чому знайшло
        ws.column_dimensions['M'].width = 30   # Оригінал
        ws.column_dimensions['N'].width = 35   # Нормалізовано
        for i, fl in enumerate(flags, start=2):
            fill = RED if fl == 'nf' else (YELLOW if fl == 'warn' else None)
            if fill:
                for cell in ws[i]:
                    cell.fill = fill

    output.seek(0)
    return output, not_found, warn

"""
logger.py — Логування та статистика.

  log_usage()        — фіксує кожне замовлення (хто, скільки, % знайдено)
  get_usage_stats()  — форматований звіт для адміна
  log_not_found()    — накопичує незнайдені позиції
  get_catalog_gaps() — топ незнайдених ("діри каталогу")
"""

import os
import json
import time

DATA_DIR       = os.environ.get("DATA_DIR") or ("/var/data" if os.path.isdir("/var/data") else ".")
USAGE_LOG_FILE = os.path.join(DATA_DIR, "usage_log.json")   # лог кожного замовлення
NOT_FOUND_FILE = os.path.join(DATA_DIR, "not_found_log.json")  # лог незнайдених позицій


# ─── Статистика використання ─────────────────────────────────────────────────

def log_usage(chat_id: int, username: str, total: int, found: int, files: int):  # записує одне замовлення в лог (chat, дата, к-ть позицій, знайдено)
    try:
        log = []
        if os.path.exists(USAGE_LOG_FILE):
            with open(USAGE_LOG_FILE, encoding="utf-8") as f:
                log = json.load(f)
        log.append({
            "chat_id":  chat_id,
            "username": username,
            "date":     time.strftime("%Y-%m-%d %H:%M"),
            "total":    total,
            "found":    found,
            "files":    files,
        })
        log = log[-1000:]   # тримаємо не більше 1000 останніх записів
        with open(USAGE_LOG_FILE, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False)
    except Exception as e:
        print(f"⚠️ log_usage: {e}")


def get_usage_stats() -> str:   # повертає форматований текстовий звіт по користувачах для адміна
    if not os.path.exists(USAGE_LOG_FILE):
        return "📊 Статистика порожня."
    try:
        with open(USAGE_LOG_FILE, encoding="utf-8") as f:
            log = json.load(f)
    except Exception:
        return "⚠️ Помилка читання логу."
    if not log:
        return "📊 Статистика порожня."

    users = {}
    for rec in log:
        u = rec.get("username", "?")
        if u not in users:
            users[u] = {"orders": 0, "total": 0, "found": 0}
        users[u]["orders"] += 1
        users[u]["total"]  += rec.get("total", 0)
        users[u]["found"]  += rec.get("found", 0)

    lines = [f"📊 *Статистика* ({len(log)} замовлень)\n"]
    for u, s in sorted(users.items(), key=lambda x: -x[1]["orders"]):
        pct = int(s["found"] / s["total"] * 100) if s["total"] else 0
        lines.append(f"👤 {u}: {s['orders']} замовл., {s['total']} поз., {pct}% знайдено")
    lines.append("\n🕐 Останні 5:")
    for rec in log[-5:]:
        lines.append(f"• {rec['date']} — {rec['username']}: {rec['found']}/{rec['total']}")
    return "\n".join(lines)


# ─── Діри каталогу ───────────────────────────────────────────────────────────

def log_not_found(rows: list):  # дописує незнайдені позиції замовлення в окремий лог для аналізу
    if not rows:
        return
    import re
    try:
        log = []
        if os.path.exists(NOT_FOUND_FILE):
            with open(NOT_FOUND_FILE, encoding="utf-8") as f:
                log = json.load(f)
        for r in rows:
            log.append({
                "original":   r.get("original", "")[:80],
                "normalized": r.get("normalized", "")[:80],
                "date":       time.strftime("%Y-%m-%d"),
            })
        log = log[-2000:]   # тримаємо не більше 2000 останніх незнайдених
        with open(NOT_FOUND_FILE, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False)
    except Exception as e:
        print(f"⚠️ log_not_found: {e}")


def get_catalog_gaps_excel(output_path: str) -> int:    # будує Excel з усіма незнайденими позиціями згрупованими по частоті; повертає кількість унікальних позицій
    import re
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    if not os.path.exists(NOT_FOUND_FILE):
        return 0
    try:
        with open(NOT_FOUND_FILE, encoding="utf-8") as f:
            log = json.load(f)
    except Exception:
        return 0
    if not log:
        return 0

    # Групуємо по normalized (або original)
    groups = {}
    for rec in log:
        key = re.sub(r"\s+", " ",
                     (rec.get("normalized") or rec.get("original", "")).lower()).strip()
        if not key:
            continue
        if key not in groups:
            groups[key] = {
                "count":      0,
                "normalized": rec.get("normalized", ""),
                "original":   rec.get("original", ""),
                "last_date":  rec.get("date", ""),
                "dates":      [],
            }
        groups[key]["count"] += 1
        groups[key]["dates"].append(rec.get("date", ""))
        if rec.get("date", "") > groups[key]["last_date"]:
            groups[key]["last_date"] = rec.get("date", "")

    # Сортуємо по частоті
    sorted_groups = sorted(groups.values(), key=lambda g: -g["count"])

    # Будуємо Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Діри каталогу"

    # Заголовки
    headers = ["№", "К-сть", "Нормалізовано (що шукав бот)", "Оригінал (що писав майстер)",
               "Остання дата", "Дія (заповнити вручну)"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font        = header_font
        cell.fill        = header_fill
        cell.alignment   = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # Дані
    red_fill    = PatternFill("solid", fgColor="FFE0E0")   # >5 разів — критично
    yellow_fill = PatternFill("solid", fgColor="FFFACD")   # 2-5 разів
    normal_font = Font(name="Arial", size=10)

    for ri, g in enumerate(sorted_groups, 2):
        row_fill = red_fill if g["count"] >= 5 else (yellow_fill if g["count"] >= 2 else None)
        values   = [
            ri - 1,
            g["count"],
            g["normalized"] or "",
            g["original"] or "",
            g["last_date"] or "",
            "",   # Дія — порожня для заповнення
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = normal_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_fill:
                cell.fill = row_fill

    # Ширина колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 30

    # Заморожуємо шапку
    ws.freeze_panes = "A2"

    # Автофільтр
    ws.auto_filter.ref = f"A1:F{len(sorted_groups)+1}"

    wb.save(output_path)
    return len(sorted_groups)


    import re
    if not os.path.exists(NOT_FOUND_FILE):
        return "🕳 Порожньо — все знаходилось."
    try:
        with open(NOT_FOUND_FILE, encoding="utf-8") as f:
            log = json.load(f)
    except Exception:
        return "⚠️ Помилка читання."
    if not log:
        return "🕳 Порожньо."

    groups = {}
    for rec in log:
        key = re.sub(r"\s+", " ",
                     (rec.get("normalized") or rec.get("original", "")).lower()).strip()
        if not key:
            continue
        if key not in groups:
            groups[key] = {
                "n":    0,
                "show": rec.get("normalized") or rec.get("original", ""),
                "orig": rec.get("original", ""),
                "last": rec.get("date", ""),
            }
        groups[key]["n"] += 1
        groups[key]["last"] = rec.get("date", groups[key]["last"])

    top   = sorted(groups.values(), key=lambda g: -g["n"])[:20]
    lines = [f"🕳 ДІРИ КАТАЛОГУ — топ незнайдених ({len(log)} записів):\n"]
    for i, g in enumerate(top, 1):
        lines.append(f"{i}. ×{g['n']}  {g['show'][:48]}")
        if g['orig'] and g['orig'].lower() != g['show'].lower():
            lines.append(f"      (писали: {g['orig'][:45]})")
    lines.append("\n➡️ Ці товари варто додати в прайси або створити правило.")
    return "\n".join(lines)
